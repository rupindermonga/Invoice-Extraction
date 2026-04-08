"""File tools: Invoice Finder (CSV → search folders) and Bulk Folder Upload."""
import os
import shutil
import csv
import io
import logging
from pathlib import Path
from typing import List, Optional
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, BackgroundTasks, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from pydantic import BaseModel

from ..database import get_db
from ..models import Invoice, User
from ..dependencies import get_current_user
from ..services.extractor import save_upload_file, process_invoice_file
from ..services.gemini import check_api_key
from .invoices import processing_store

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/filetools", tags=["filetools"])

ALLOWED_EXTENSIONS = {".pdf", ".jpg", ".jpeg", ".png", ".tiff", ".tif", ".webp"}


def _extract_text_from_pdf(filepath: str) -> str:
    """Extract text from a PDF using PyPDF2 (free, no Gemini cost)."""
    try:
        from PyPDF2 import PdfReader
        reader = PdfReader(filepath)
        text = ""
        for page in reader.pages[:5]:  # Only check first 5 pages
            text += (page.extract_text() or "") + "\n"
        return text
    except Exception as e:
        logger.warning("Could not extract text from %s: %s", filepath, e)
        return ""


def _search_folder_filename_only(folder: str, invoice_number: str) -> Optional[str]:
    """Fast search: filename match only, no PDF reading."""
    if not os.path.isdir(folder):
        return None
    inv_num_lower = invoice_number.lower().strip()

    # Direct files
    for entry in os.scandir(folder):
        if entry.is_file() and Path(entry.name).suffix.lower() in ALLOWED_EXTENSIONS:
            if inv_num_lower in entry.name.lower():
                return entry.path

    # Subfolders
    for entry in os.scandir(folder):
        if entry.is_dir():
            for sub_entry in os.scandir(entry.path):
                if sub_entry.is_file() and Path(sub_entry.name).suffix.lower() in ALLOWED_EXTENSIONS:
                    if inv_num_lower in sub_entry.name.lower():
                        return sub_entry.path
    return None


# Cache for PDF text extraction (avoids re-reading same file)
_pdf_text_cache: dict = {}


def _search_folder_deep(folder: str, invoice_number: str) -> Optional[str]:
    """Deep search: reads PDF text content. Uses cache to avoid re-reading."""
    if not os.path.isdir(folder):
        return None
    inv_num_lower = invoice_number.lower().strip()

    # Direct PDF files
    for entry in os.scandir(folder):
        if entry.is_file() and entry.name.lower().endswith(".pdf"):
            if entry.path not in _pdf_text_cache:
                _pdf_text_cache[entry.path] = _extract_text_from_pdf(entry.path)
            if inv_num_lower in _pdf_text_cache[entry.path].lower():
                return entry.path

    # Subfolder PDFs
    for entry in os.scandir(folder):
        if entry.is_dir():
            for sub_entry in os.scandir(entry.path):
                if sub_entry.is_file() and sub_entry.name.lower().endswith(".pdf"):
                    if sub_entry.path not in _pdf_text_cache:
                        _pdf_text_cache[sub_entry.path] = _extract_text_from_pdf(sub_entry.path)
                    if inv_num_lower in _pdf_text_cache[sub_entry.path].lower():
                        return sub_entry.path
    return None


def _find_vendor_folder(source: str, vendor: str) -> Optional[str]:
    """Find the best matching vendor subfolder using fuzzy matching.
    Tries: exact match, then case-insensitive, then partial (vendor name contains or is contained)."""
    if not os.path.isdir(source):
        return None
    vendor_lower = vendor.lower().strip()
    # Extract first word / base name (e.g. "Valard" from "Valard - Hydro")
    vendor_base = vendor_lower.split("-")[0].split("–")[0].strip()

    exact = os.path.join(source, vendor)
    if os.path.isdir(exact):
        return exact

    # Scan all subdirs for best match
    best = None
    for entry in os.scandir(source):
        if not entry.is_dir():
            continue
        folder_lower = entry.name.lower().strip()
        # Case-insensitive exact
        if folder_lower == vendor_lower:
            return entry.path
        # Folder contains vendor base or vendor base contains folder name
        if vendor_base and (vendor_base in folder_lower or folder_lower in vendor_base):
            best = entry.path
        # Also check if any word in vendor matches folder
        elif not best:
            vendor_words = [w.strip() for w in vendor_lower.replace("-", " ").replace("–", " ").split() if len(w.strip()) > 2]
            for w in vendor_words:
                if w in folder_lower:
                    best = entry.path
                    break
    return best


class InvoiceFinderRequest(BaseModel):
    source_folder: str
    output_folder: Optional[str] = None
    invoices: List[dict]  # [{"vendor": "Nokia", "invoice_number": "123"}, ...]
    mode: str = "fast"  # "fast" = filename only, "deep" = also read PDF content


# Global cancel flag for long-running searches
_cancel_search: dict = {}  # user_id -> bool


@router.post("/find-invoices/cancel")
def cancel_search(current_user: User = Depends(get_current_user)):
    """Cancel an in-progress invoice search."""
    _cancel_search[current_user.id] = True
    return {"message": "Cancel requested"}


@router.post("/find-invoices")
def find_invoices(
    body: InvoiceFinderRequest,
    current_user: User = Depends(get_current_user),
):
    """Search source folder for invoices. mode=fast (filename only) or mode=deep (reads PDFs too)."""
    import json as _json
    _cancel_search[current_user.id] = False
    source = body.source_folder.strip()
    if not os.path.isdir(source):
        raise HTTPException(status_code=400, detail=f"Source folder not found: {source}")

    output = (body.output_folder or "").strip()
    if not output:
        output = os.path.join(source, "_organized")
    os.makedirs(output, exist_ok=True)

    deep = body.mode == "deep"
    found = []
    missing = []
    duplicates = []
    cancelled = False
    total = len(body.invoices)

    for idx, item in enumerate(body.invoices):
        # Check cancel flag
        if _cancel_search.get(current_user.id):
            cancelled = True
            break

        vendor = (item.get("vendor") or "Unknown").strip()
        inv_num = (item.get("invoice_number") or "").strip()
        if not inv_num:
            missing.append({"vendor": vendor, "invoice_number": inv_num, "reason": "No invoice number"})
            continue

        # Find best matching vendor folder (fuzzy)
        vendor_folder = _find_vendor_folder(source, vendor)
        match = None

        # Fast pass: filename search
        if vendor_folder:
            match = _search_folder_filename_only(vendor_folder, inv_num)
        if not match:
            match = _search_folder_filename_only(source, inv_num)
        if not match:
            for entry in os.scandir(source):
                if entry.is_dir() and entry.path != vendor_folder:
                    match = _search_folder_filename_only(entry.path, inv_num)
                    if match:
                        break

        # Deep pass: read PDF text (only if mode=deep and still not found)
        if not match and deep:
            if vendor_folder:
                match = _search_folder_deep(vendor_folder, inv_num)
            if not match:
                match = _search_folder_deep(source, inv_num)
            if not match:
                for entry in os.scandir(source):
                    if entry.is_dir() and entry.path != vendor_folder:
                        match = _search_folder_deep(entry.path, inv_num)
                        if match:
                            break

        if match:
            dest_dir = os.path.join(output, vendor)
            os.makedirs(dest_dir, exist_ok=True)
            dest_file = os.path.join(dest_dir, os.path.basename(match))
            already_exists = os.path.exists(dest_file)
            if not already_exists:
                shutil.copy2(match, dest_file)
            entry_data = {
                "vendor": vendor,
                "invoice_number": inv_num,
                "source_path": match,
                "dest_path": dest_file,
            }
            if already_exists:
                duplicates.append(entry_data)
            else:
                found.append(entry_data)
        else:
            missing.append({
                "vendor": vendor,
                "invoice_number": inv_num,
                "reason": "Not found in source folder",
            })

    _cancel_search.pop(current_user.id, None)

    return {
        "cancelled": cancelled,
        "total": total,
        "searched": idx + 1 if not cancelled else idx,
        "found": len(found),
        "duplicates": len(duplicates),
        "missing": len(missing),
        "found_list": found,
        "duplicate_list": duplicates,
        "missing_list": missing,
        "output_folder": output,
    }


@router.post("/upload-csv")
async def upload_invoice_csv(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
):
    """Parse a CSV/Excel file with Vendor, Invoice Number columns. Returns the parsed list."""
    content = await file.read()
    name = (file.filename or "").lower()

    rows = []
    if name.endswith(".csv"):
        text = content.decode("utf-8-sig")
        reader = csv.reader(io.StringIO(text))
        all_rows = list(reader)
        if not all_rows:
            raise HTTPException(status_code=400, detail="Empty CSV")
        # Find first two non-empty columns from header row
        header = all_rows[0]
        data_cols = [i for i, v in enumerate(header) if v and v.strip()]
        if len(data_cols) < 2:
            # Fallback to first two columns
            data_cols = [0, 1]
        vc, ic = data_cols[0], data_cols[1]
        for row in all_rows[1:]:  # skip header
            if len(row) <= max(vc, ic):
                continue
            vendor = str(row[vc]).strip()
            inv_num = str(row[ic]).strip()
            if vendor or inv_num:
                rows.append({"vendor": vendor, "invoice_number": inv_num})
    elif name.endswith((".xlsx", ".xls")):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
        ws = wb.active
        # Find first two non-empty columns from header row
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            raise HTTPException(status_code=400, detail="Empty spreadsheet")
        # Find column indices that have data (skip empty columns)
        data_cols = [i for i, v in enumerate(header_row) if v is not None and str(v).strip()]
        if len(data_cols) < 2:
            raise HTTPException(status_code=400, detail="Need at least 2 columns with data")
        vendor_col = data_cols[0]
        inv_col = data_cols[1]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row is None:
                continue
            vendor = str(row[vendor_col] if vendor_col < len(row) else "" or "").strip()
            inv_num = str(row[inv_col] if inv_col < len(row) else "" or "").strip()
            if vendor or inv_num:
                rows.append({"vendor": vendor, "invoice_number": inv_num})
    else:
        raise HTTPException(status_code=400, detail="Unsupported file type. Upload CSV or Excel (.xlsx)")

    return {"count": len(rows), "invoices": rows}


@router.post("/bulk-upload-folder")
async def bulk_upload_folder(
    background_tasks: BackgroundTasks,
    folder_path: str = Query(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Scan a local folder recursively, upload all invoice files, and extract with Gemini."""
    folder = folder_path.strip()
    if not os.path.isdir(folder):
        raise HTTPException(status_code=400, detail=f"Folder not found: {folder}")

    if not check_api_key(db):
        raise HTTPException(status_code=400, detail="No Gemini API key configured.")

    results = []
    for root, dirs, files in os.walk(folder):
        for fname in sorted(files):
            ext = Path(fname).suffix.lower()
            if ext not in ALLOWED_EXTENSIONS:
                continue

            filepath = os.path.join(root, fname)
            # Skip files already uploaded (by filename)
            existing = db.query(Invoice).filter(
                Invoice.user_id == current_user.id,
                Invoice.original_filename == fname,
            ).first()
            if existing:
                results.append({"filename": fname, "status": "skipped", "reason": "Already uploaded"})
                continue

            # Read and save
            try:
                with open(filepath, "rb") as f:
                    content = f.read()
            except Exception as e:
                results.append({"filename": fname, "status": "error", "reason": str(e)})
                continue

            saved_path = save_upload_file(content, fname)
            invoice = Invoice(
                user_id=current_user.id,
                source="folder",
                source_file=saved_path,
                original_filename=fname,
                status="pending",
            )
            db.add(invoice)
            db.commit()
            db.refresh(invoice)

            background_tasks.add_task(
                process_invoice_file,
                invoice.id,
                saved_path,
                current_user.id,
                db,
                processing_store,
            )
            results.append({"invoice_id": invoice.id, "filename": fname, "status": "queued"})

    return {"total": len(results), "results": results}
