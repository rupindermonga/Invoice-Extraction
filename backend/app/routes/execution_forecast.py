"""
Execution Forecast — Generic Construction Progress Tracking

Works for any infrastructure project (fibre, roads, utilities, etc.):
  - Workstreams  : phases of work (Design, Permits, Build, Test…)
  - Nodes        : work zones (OLTs, Sections, Zones…)
  - Tasks        : one row per Node × Workstream (the schedule)
  - Progress     : planned + actual quantities per period (monthly / weekly / daily)

Entry methods:
  1. Excel import  — reads the standard template (OLT | SD | Workstream | Vendor |
                     Start Date | Deadline | Unit | Total Scope | Remaining |
                     [Month Plan, Month Actual]…)
  2. Manual CRUD   — add nodes, workstreams, tasks one at a time
  3. Bulk actuals  — PM selects a period and enters actuals for all tasks at once
"""
from __future__ import annotations

import io
import json
import re
from datetime import datetime
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import text

from ..database import SessionLocal
from ..dependencies import (
    get_current_user, require_org_member, require_project_access,
    FINANCE_READ_ROLES, FINANCE_WRITE_ROLES,
)
from ..models import User

router = APIRouter(prefix="/api/project", tags=["execution-forecast"])


def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


# ─── Helpers ──────────────────────────────────────────────────────────────────

MONTH_ABBR = {
    "jan": "01", "feb": "02", "mar": "03", "apr": "04",
    "may": "05", "jun": "06", "jul": "07", "aug": "08",
    "sep": "09", "oct": "10", "nov": "11", "dec": "12",
}


def _parse_period(col_header: str) -> Optional[str]:
    """Convert 'Apr-2026 Plan' → ('2026-04', 'plan') or None."""
    h = col_header.strip().lower()
    m = re.match(r"([a-z]{3})-(\d{4})\s+(plan|actual)", h)
    if not m:
        return None
    mon, year, kind = m.group(1), m.group(2), m.group(3)
    if mon not in MONTH_ABBR:
        return None
    return f"{year}-{MONTH_ABBR[mon]}", kind


def _period_label(iso: str) -> str:
    """'2026-04' → 'Apr-2026'"""
    rev = {v: k.capitalize() for k, v in MONTH_ABBR.items()}
    try:
        y, m = iso.split("-")
        return f"{rev[m]}-{y}"
    except Exception:
        return iso


def _node_key(node_code: str, sd_code: str) -> str:
    return f"{sd_code}|{node_code}"


# ─── Workstreams CRUD ─────────────────────────────────────────────────────────

@router.get("/{project_id}/execution/workstreams")
def list_workstreams(project_id: int, db: Session = Depends(get_db),
                     current_user: User = Depends(get_current_user)):
    require_org_member(db, current_user.org_id, current_user.id, FINANCE_READ_ROLES)
    require_project_access(db, project_id, current_user.org_id)
    rows = db.execute(text(
        "SELECT id, name, unit, display_order FROM wf_workstreams "
        "WHERE project_id=:pid AND org_id=:oid ORDER BY display_order, id"
    ), {"pid": project_id, "oid": current_user.org_id}).fetchall()
    return [{"id": r[0], "name": r[1], "unit": r[2], "display_order": r[3]} for r in rows]


@router.post("/{project_id}/execution/workstreams")
def create_workstream(project_id: int, body: dict,
                      db: Session = Depends(get_db),
                      current_user: User = Depends(get_current_user)):
    require_org_member(db, current_user.org_id, current_user.id, FINANCE_WRITE_ROLES)
    require_project_access(db, project_id, current_user.org_id)
    name = (body.get("name") or "").strip()
    if not name:
        raise HTTPException(400, "name required")
    r = db.execute(text(
        "INSERT INTO wf_workstreams (org_id, project_id, name, unit, display_order) "
        "VALUES (:oid, :pid, :name, :unit, :ord) RETURNING id"
    ), {"oid": current_user.org_id, "pid": project_id,
        "name": name, "unit": body.get("unit", "units"),
        "ord": body.get("display_order", 100)})
    db.commit()
    return {"id": r.fetchone()[0], "name": name}


@router.delete("/{project_id}/execution/workstreams/{ws_id}")
def delete_workstream(project_id: int, ws_id: int,
                      db: Session = Depends(get_db),
                      current_user: User = Depends(get_current_user)):
    require_org_member(db, current_user.org_id, current_user.id, FINANCE_WRITE_ROLES)
    require_project_access(db, project_id, current_user.org_id)
    db.execute(text(
        "DELETE FROM wf_workstreams WHERE id=:id AND project_id=:pid AND org_id=:oid"
    ), {"id": ws_id, "pid": project_id, "oid": current_user.org_id})
    db.commit()
    return {"msg": "deleted"}


# ─── Nodes CRUD ───────────────────────────────────────────────────────────────

@router.get("/{project_id}/execution/nodes")
def list_nodes(project_id: int, db: Session = Depends(get_db),
               current_user: User = Depends(get_current_user)):
    require_org_member(db, current_user.org_id, current_user.id, FINANCE_READ_ROLES)
    require_project_access(db, project_id, current_user.org_id)
    rows = db.execute(text(
        "SELECT id, sd_code, node_code, node_name, vendor, display_order "
        "FROM wf_nodes WHERE project_id=:pid AND org_id=:oid "
        "ORDER BY sd_code, display_order, node_code"
    ), {"pid": project_id, "oid": current_user.org_id}).fetchall()
    return [{"id": r[0], "sd_code": r[1], "node_code": r[2],
             "node_name": r[3], "vendor": r[4], "display_order": r[5]} for r in rows]


@router.post("/{project_id}/execution/nodes")
def create_node(project_id: int, body: dict,
                db: Session = Depends(get_db),
                current_user: User = Depends(get_current_user)):
    require_org_member(db, current_user.org_id, current_user.id, FINANCE_WRITE_ROLES)
    require_project_access(db, project_id, current_user.org_id)
    code = (body.get("node_code") or "").strip()
    if not code:
        raise HTTPException(400, "node_code required")
    r = db.execute(text(
        "INSERT INTO wf_nodes (org_id, project_id, sd_code, node_code, node_name, vendor, display_order) "
        "VALUES (:oid, :pid, :sd, :code, :name, :vendor, :ord) RETURNING id"
    ), {"oid": current_user.org_id, "pid": project_id,
        "sd": body.get("sd_code"), "code": code,
        "name": body.get("node_name"), "vendor": body.get("vendor"),
        "ord": body.get("display_order", 100)})
    db.commit()
    return {"id": r.fetchone()[0], "node_code": code}


@router.delete("/{project_id}/execution/nodes/{node_id}")
def delete_node(project_id: int, node_id: int,
                db: Session = Depends(get_db),
                current_user: User = Depends(get_current_user)):
    require_org_member(db, current_user.org_id, current_user.id, FINANCE_WRITE_ROLES)
    require_project_access(db, project_id, current_user.org_id)
    db.execute(text(
        "DELETE FROM wf_nodes WHERE id=:id AND project_id=:pid AND org_id=:oid"
    ), {"id": node_id, "pid": project_id, "oid": current_user.org_id})
    db.commit()
    return {"msg": "deleted"}


# ─── Excel Import ─────────────────────────────────────────────────────────────

@router.post("/{project_id}/execution/import")
async def import_excel(project_id: int,
                       file: UploadFile = File(...),
                       db: Session = Depends(get_db),
                       current_user: User = Depends(get_current_user)):
    """
    Import an execution forecast from Excel.

    Expected sheet name: 'Detailed Project Plan' (falls back to first sheet).
    Expected columns: OLT | SD | Workstream | Vendor | Start Date | Deadline |
                      Unit | Total Scope | Remaining | [Mon-YYYY Plan, Mon-YYYY Actual]…

    Nodes and workstreams are auto-created if they don't exist.
    Existing planned quantities are overwritten; existing actuals are preserved.
    """
    require_org_member(db, current_user.org_id, current_user.id, FINANCE_WRITE_ROLES)
    require_project_access(db, project_id, current_user.org_id)

    content = await file.read()
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(400, f"Cannot open Excel file: {e}")

    ws = None
    for sname in wb.sheetnames:          # prefer "detailed" first
        if "detailed" in sname.lower():
            ws = wb[sname]; break
    if ws is None:
        for sname in wb.sheetnames:      # fallback: any sheet with "plan"
            if "plan" in sname.lower():
                ws = wb[sname]; break
    if ws is None:
        ws = wb[wb.sheetnames[0]]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(400, "Sheet is empty")

    header = [str(c).strip() if c is not None else "" for c in rows[0]]

    # Locate fixed columns
    def _col(name):
        nl = name.lower()
        for i, h in enumerate(header):
            if h.lower() == nl:
                return i
        return None

    ci = {
        "olt":        _col("olt") if _col("olt") is not None else 0,
        "sd":         _col("sd")  if _col("sd")  is not None else 1,
        "workstream": next((i for i, h in enumerate(header) if "workstream" in h.lower()), 2),
        "vendor":     next((i for i, h in enumerate(header) if "vendor" in h.lower()), 3),
        "start":      next((i for i, h in enumerate(header) if "start" in h.lower()), 4),
        "deadline":   next((i for i, h in enumerate(header) if "deadline" in h.lower()), 5),
        "unit":       next((i for i, h in enumerate(header) if h.lower() == "unit"), 6),
        "scope":      next((i for i, h in enumerate(header) if "scope" in h.lower()), 7),
        "remaining":  next((i for i, h in enumerate(header) if "remaining" in h.lower()), 8),
    }

    # Parse period columns
    period_cols = []  # list of (col_index, iso_period, 'plan'|'actual')
    for i, h in enumerate(header):
        if i <= ci["remaining"]:
            continue
        parsed = _parse_period(h)
        if parsed:
            period_cols.append((i, parsed[0], parsed[1]))

    if not period_cols:
        raise HTTPException(400, "No period columns found. Expected headers like 'Apr-2026 Plan', 'Apr-2026 Actual'.")

    oid = current_user.org_id
    pid = project_id

    # Build lookup caches to avoid repeated queries
    ws_cache: dict[str, int] = {}   # name → id
    node_cache: dict[str, int] = {} # "SD|OLT" → id
    task_cache: dict[str, int] = {} # "node_id|ws_id" → id

    def _get_or_create_workstream(name: str, unit: str) -> int:
        if name in ws_cache:
            return ws_cache[name]
        row = db.execute(text(
            "SELECT id FROM wf_workstreams WHERE project_id=:pid AND org_id=:oid AND name=:name"
        ), {"pid": pid, "oid": oid, "name": name}).fetchone()
        if row:
            ws_cache[name] = row[0]
            return row[0]
        r = db.execute(text(
            "INSERT INTO wf_workstreams (org_id, project_id, name, unit, display_order) "
            "VALUES (:oid, :pid, :name, :unit, :ord) RETURNING id"
        ), {"oid": oid, "pid": pid, "name": name, "unit": unit or "units", "ord": len(ws_cache) * 10})
        db.flush()
        new_id = r.fetchone()[0]
        ws_cache[name] = new_id
        return new_id

    def _get_or_create_node(node_code: str, sd_code: str, vendor: str) -> int:
        key = _node_key(node_code, sd_code or "")
        if key in node_cache:
            return node_cache[key]
        row = db.execute(text(
            "SELECT id FROM wf_nodes WHERE project_id=:pid AND org_id=:oid "
            "AND node_code=:nc AND (sd_code=:sd OR (sd_code IS NULL AND :sd IS NULL))"
        ), {"pid": pid, "oid": oid, "nc": node_code, "sd": sd_code or None}).fetchone()
        if row:
            node_cache[key] = row[0]
            return row[0]
        r = db.execute(text(
            "INSERT INTO wf_nodes (org_id, project_id, sd_code, node_code, vendor, display_order) "
            "VALUES (:oid, :pid, :sd, :nc, :vendor, :ord) RETURNING id"
        ), {"oid": oid, "pid": pid, "sd": sd_code or None, "nc": node_code,
            "vendor": vendor or None, "ord": len(node_cache) * 10})
        db.flush()
        new_id = r.fetchone()[0]
        node_cache[key] = new_id
        return new_id

    def _get_or_create_task(node_id: int, ws_id: int,
                            vendor: str, start: str, deadline: str,
                            scope: float, remaining: float) -> int:
        key = f"{node_id}|{ws_id}"
        if key in task_cache:
            return task_cache[key]
        row = db.execute(text(
            "SELECT id FROM wf_tasks WHERE node_id=:nid AND workstream_id=:wid AND project_id=:pid"
        ), {"nid": node_id, "wid": ws_id, "pid": pid}).fetchone()
        if row:
            task_cache[key] = row[0]
            return row[0]
        r = db.execute(text(
            "INSERT INTO wf_tasks (org_id, project_id, node_id, workstream_id, vendor, "
            "start_date, deadline, total_scope, remaining) "
            "VALUES (:oid, :pid, :nid, :wid, :vendor, :start, :deadline, :scope, :rem) RETURNING id"
        ), {"oid": oid, "pid": pid, "nid": node_id, "wid": ws_id,
            "vendor": vendor or None, "start": str(start) if start else None,
            "deadline": str(deadline) if deadline else None,
            "scope": float(scope or 0), "rem": float(remaining or 0)})
        db.flush()
        new_id = r.fetchone()[0]
        task_cache[key] = new_id
        return new_id

    stats = {"rows_processed": 0, "nodes_created": 0, "workstreams_created": 0,
             "tasks_created": 0, "plan_cells": 0, "actual_cells": 0}
    prev_ws_count = 0
    prev_node_count = 0

    for row in rows[1:]:
        if not row or all(v is None for v in row):
            continue
        node_code = str(row[ci["olt"]]).strip() if row[ci["olt"]] else None
        sd_code   = str(row[ci["sd"]]).strip()  if row[ci["sd"]]  else None
        ws_name   = str(row[ci["workstream"]]).strip() if row[ci["workstream"]] else None
        if not node_code or not ws_name:
            continue

        vendor   = str(row[ci["vendor"]]).strip()   if row[ci["vendor"]]   else None
        start    = row[ci["start"]]
        deadline = row[ci["deadline"]]
        unit     = str(row[ci["unit"]]).strip()     if row[ci["unit"]]     else "units"
        scope    = row[ci["scope"]]    or 0
        remaining = row[ci["remaining"]] or 0

        ws_id   = _get_or_create_workstream(ws_name, unit)
        node_id = _get_or_create_node(node_code, sd_code, vendor)
        task_id = _get_or_create_task(node_id, ws_id, vendor, start, deadline, scope, remaining)

        stats["workstreams_created"] += len(ws_cache) - prev_ws_count
        stats["nodes_created"] += len(node_cache) - prev_node_count
        prev_ws_count = len(ws_cache)
        prev_node_count = len(node_cache)

        # Upsert progress rows
        plan_vals: dict[str, float] = {}
        actual_vals: dict[str, float] = {}
        for col_idx, iso_period, kind in period_cols:
            val = row[col_idx] if col_idx < len(row) else None
            if val is not None:
                try:
                    fval = float(val)
                except (ValueError, TypeError):
                    continue
                if kind == "plan":
                    plan_vals[iso_period] = fval
                else:
                    actual_vals[iso_period] = fval

        all_periods = set(plan_vals) | set(actual_vals)
        for period in all_periods:
            planned = plan_vals.get(period, 0)
            actual  = actual_vals.get(period)
            existing = db.execute(text(
                "SELECT id, actual_qty FROM wf_progress WHERE task_id=:tid AND period=:p"
            ), {"tid": task_id, "p": period}).fetchone()
            if existing:
                # Overwrite plan; preserve existing actual unless file has one
                new_actual = actual if actual is not None else existing[1]
                db.execute(text(
                    "UPDATE wf_progress SET planned_qty=:plan, actual_qty=:act WHERE id=:id"
                ), {"plan": planned, "act": new_actual, "id": existing[0]})
            else:
                db.execute(text(
                    "INSERT INTO wf_progress (task_id, org_id, project_id, period, period_type, "
                    "planned_qty, actual_qty) VALUES (:tid, :oid, :pid, :period, 'monthly', :plan, :act)"
                ), {"tid": task_id, "oid": oid, "pid": pid, "period": period,
                    "plan": planned, "act": actual})
            if kind == "plan" or planned:
                stats["plan_cells"] += 1
            if actual is not None:
                stats["actual_cells"] += 1

        stats["rows_processed"] += 1

    stats["tasks_created"] = len(task_cache)
    db.commit()
    return {
        "msg": "Import complete",
        "sheet_used": ws.title,
        "periods_found": len(set(p for _, p, _ in period_cols)),
        **stats,
    }


# ─── Grid: full plan vs actual table ─────────────────────────────────────────

@router.get("/{project_id}/execution/grid")
def get_grid(project_id: int,
             sd_code: Optional[str] = None,
             node_id: Optional[int] = None,
             node_code: Optional[str] = None,
             vendor: Optional[str] = None,
             workstream_id: Optional[int] = None,
             db: Session = Depends(get_db),
             current_user: User = Depends(get_current_user)):
    """
    Returns a flat list of task rows, each with all period plan/actual values.
    Filter by sd_code, node_id, node_code, vendor, or workstream_id.
    """
    require_org_member(db, current_user.org_id, current_user.id, FINANCE_READ_ROLES)
    require_project_access(db, project_id, current_user.org_id)
    oid = current_user.org_id
    pid = project_id

    where = "t.project_id=:pid AND t.org_id=:oid"
    params: dict = {"pid": pid, "oid": oid}
    if sd_code:
        where += " AND n.sd_code=:sd"
        params["sd"] = sd_code
    if node_id:
        where += " AND t.node_id=:nid"
        params["nid"] = node_id
    if workstream_id:
        where += " AND t.workstream_id=:wsid"
        params["wsid"] = workstream_id
    if node_code:
        where += " AND n.node_code=:nc"
        params["nc"] = node_code
    if vendor:
        where += " AND t.vendor=:vendor"
        params["vendor"] = vendor

    tasks = db.execute(text(f"""
        SELECT t.id, n.sd_code, n.node_code, w.name AS workstream, w.unit,
               t.vendor, t.start_date, t.deadline, t.total_scope, t.remaining,
               w.display_order AS ws_order, n.display_order AS node_order
        FROM wf_tasks t
        JOIN wf_nodes n ON n.id = t.node_id
        JOIN wf_workstreams w ON w.id = t.workstream_id
        WHERE {where}
        ORDER BY n.sd_code, n.display_order, n.node_code, w.display_order
    """), params).fetchall()

    if not tasks:
        return {"tasks": [], "periods": []}

    task_ids = [r[0] for r in tasks]
    progress = db.execute(text(
        f"SELECT task_id, period, planned_qty, actual_qty FROM wf_progress "
        f"WHERE task_id IN ({','.join(str(i) for i in task_ids)}) "
        f"ORDER BY period"
    )).fetchall()

    # Build period → plan/actual map per task
    prog_map: dict[int, dict[str, dict]] = {}
    all_periods: set[str] = set()
    for tid, period, plan, actual in progress:
        prog_map.setdefault(tid, {})[period] = {"plan": plan or 0, "actual": actual}
        all_periods.add(period)

    periods_sorted = sorted(all_periods)

    result = []
    for t in tasks:
        tid = t[0]
        periods_data = prog_map.get(tid, {})
        result.append({
            "task_id": tid,
            "sd_code": t[1],
            "node_code": t[2],
            "workstream": t[3],
            "unit": t[4],
            "vendor": t[5],
            "start_date": t[6],
            "deadline": t[7],
            "total_scope": t[8],
            "remaining": t[9],
            "periods": {p: periods_data.get(p, {"plan": 0, "actual": None})
                        for p in periods_sorted},
        })

    return {
        "tasks": result,
        "periods": [{"iso": p, "label": _period_label(p)} for p in periods_sorted],
    }


# ─── Summary: rolled up by workstream ────────────────────────────────────────

@router.get("/{project_id}/execution/summary")
def get_summary(project_id: int,
                sd_code: Optional[str] = None,
                node_code: Optional[str] = None,
                vendor: Optional[str] = None,
                group_by: str = "workstream",
                db: Session = Depends(get_db),
                current_user: User = Depends(get_current_user)):
    """
    Returns totals rolled up by the chosen dimension.
    group_by: workstream | sd_code | node_code | vendor
    """
    require_org_member(db, current_user.org_id, current_user.id, FINANCE_READ_ROLES)
    require_project_access(db, project_id, current_user.org_id)
    oid = current_user.org_id
    pid = project_id

    # Build WHERE
    filters = "p.project_id=:pid AND p.org_id=:oid"
    params: dict = {"pid": pid, "oid": oid}
    if sd_code:
        filters += " AND n.sd_code=:sd"; params["sd"] = sd_code
    if node_code:
        filters += " AND n.node_code=:nc"; params["nc"] = node_code
    if vendor:
        filters += " AND t.vendor=:v"; params["v"] = vendor

    # Build SELECT / GROUP BY based on group_by dimension
    if group_by == "sd_code":
        grp_select = "COALESCE(n.sd_code,'Unknown') AS grp_name, '' AS unit, COALESCE(n.sd_code,'') AS grp_ord"
        grp_col = "n.sd_code"
    elif group_by == "node_code":
        grp_select = "n.node_code AS grp_name, '' AS unit, n.node_code AS grp_ord"
        grp_col = "n.node_code"
    elif group_by == "vendor":
        grp_select = "COALESCE(t.vendor,'Unknown') AS grp_name, '' AS unit, COALESCE(t.vendor,'') AS grp_ord"
        grp_col = "COALESCE(t.vendor,'Unknown')"
    else:  # workstream (default)
        grp_select = "w.name AS grp_name, w.unit AS unit, w.display_order AS grp_ord"
        grp_col = "w.id"

    ws_join = "JOIN wf_workstreams w ON w.id = t.workstream_id" if group_by == "workstream" else "LEFT JOIN wf_workstreams w ON w.id = t.workstream_id"

    rows = db.execute(text(f"""
        SELECT {grp_select}, p.period, SUM(p.planned_qty), SUM(p.actual_qty), SUM(t.total_scope)
        FROM wf_progress p
        JOIN wf_tasks t ON t.id = p.task_id
        {ws_join}
        JOIN wf_nodes n ON n.id = t.node_id
        WHERE {filters}
        GROUP BY {grp_col}, p.period
        ORDER BY grp_ord, p.period
    """), params).fetchall()

    workstreams: dict[str, dict] = {}
    all_periods: set[str] = set()
    for grp_name, unit, grp_ord, period, plan_sum, actual_sum, scope_sum in rows:
        all_periods.add(period)
        if grp_name not in workstreams:
            workstreams[grp_name] = {
                "workstream": grp_name, "unit": unit or "km",
                "total_scope": 0, "periods": {},
            }
        workstreams[grp_name]["total_scope"] = scope_sum or 0
        workstreams[grp_name]["periods"][period] = {
            "plan": plan_sum or 0,
            "actual": actual_sum,
        }

    periods_sorted = sorted(all_periods)
    return {
        "workstreams": list(workstreams.values()),
        "periods": [{"iso": p, "label": _period_label(p)} for p in periods_sorted],
        "group_by": group_by,
    }


@router.get("/{project_id}/execution/vendors")
def list_vendors(project_id: int, db: Session = Depends(get_db),
                 current_user: User = Depends(get_current_user)):
    """Return distinct vendor names for this project."""
    require_org_member(db, current_user.org_id, current_user.id, FINANCE_READ_ROLES)
    require_project_access(db, project_id, current_user.org_id)
    rows = db.execute(text(
        "SELECT DISTINCT t.vendor FROM wf_tasks t "
        "WHERE t.project_id=:pid AND t.org_id=:oid AND t.vendor IS NOT NULL "
        "ORDER BY t.vendor"
    ), {"pid": project_id, "oid": current_user.org_id}).fetchall()
    return [r[0] for r in rows]


# ─── Available periods ────────────────────────────────────────────────────────

@router.get("/{project_id}/execution/periods")
def list_periods(project_id: int, db: Session = Depends(get_db),
                 current_user: User = Depends(get_current_user)):
    require_org_member(db, current_user.org_id, current_user.id, FINANCE_READ_ROLES)
    require_project_access(db, project_id, current_user.org_id)
    rows = db.execute(text(
        "SELECT DISTINCT period FROM wf_progress WHERE project_id=:pid AND org_id=:oid ORDER BY period"
    ), {"pid": project_id, "oid": current_user.org_id}).fetchall()
    return [{"iso": r[0], "label": _period_label(r[0])} for r in rows]


# ─── Available subdivisions ───────────────────────────────────────────────────

@router.get("/{project_id}/execution/subdivisions")
def list_subdivisions(project_id: int, db: Session = Depends(get_db),
                      current_user: User = Depends(get_current_user)):
    require_org_member(db, current_user.org_id, current_user.id, FINANCE_READ_ROLES)
    require_project_access(db, project_id, current_user.org_id)
    rows = db.execute(text(
        "SELECT DISTINCT sd_code FROM wf_nodes WHERE project_id=:pid AND org_id=:oid "
        "AND sd_code IS NOT NULL ORDER BY sd_code"
    ), {"pid": project_id, "oid": current_user.org_id}).fetchall()
    return [r[0] for r in rows]


# ─── Enter / update actuals ───────────────────────────────────────────────────

@router.post("/{project_id}/execution/actuals")
def enter_actuals(project_id: int, body: dict,
                  db: Session = Depends(get_db),
                  current_user: User = Depends(get_current_user)):
    """
    Bulk enter actuals for a period.

    Body:
      period: "2026-05"
      entries: [
        { task_id: 123, actual_qty: 45.5, notes: "..." },
        ...
      ]

    Creates progress row if it doesn't exist (plan=0). Updates actual only.
    """
    require_org_member(db, current_user.org_id, current_user.id, FINANCE_WRITE_ROLES)
    require_project_access(db, project_id, current_user.org_id)
    oid = current_user.org_id
    pid = project_id
    period = (body.get("period") or "").strip()
    if not period:
        raise HTTPException(400, "period required (e.g. '2026-05')")
    entries = body.get("entries") or []
    updated = 0
    for e in entries:
        task_id = e.get("task_id")
        actual  = e.get("actual_qty")
        notes   = e.get("notes")
        if task_id is None or actual is None:
            continue
        # Verify task belongs to this project+org
        t = db.execute(text(
            "SELECT id FROM wf_tasks WHERE id=:tid AND project_id=:pid AND org_id=:oid"
        ), {"tid": task_id, "pid": pid, "oid": oid}).fetchone()
        if not t:
            continue
        existing = db.execute(text(
            "SELECT id FROM wf_progress WHERE task_id=:tid AND period=:p"
        ), {"tid": task_id, "p": period}).fetchone()
        if existing:
            db.execute(text(
                "UPDATE wf_progress SET actual_qty=:act, notes=:notes, "
                "entered_by=:uid, entered_at=:now WHERE id=:id"
            ), {"act": float(actual), "notes": notes, "uid": current_user.id,
                "now": datetime.utcnow().isoformat(), "id": existing[0]})
        else:
            db.execute(text(
                "INSERT INTO wf_progress (task_id, org_id, project_id, period, period_type, "
                "planned_qty, actual_qty, entered_by, entered_at, notes) "
                "VALUES (:tid, :oid, :pid, :period, 'monthly', 0, :act, :uid, :now, :notes)"
            ), {"tid": task_id, "oid": oid, "pid": pid, "period": period,
                "act": float(actual), "uid": current_user.id,
                "now": datetime.utcnow().isoformat(), "notes": notes})
        updated += 1
    db.commit()
    return {"msg": f"{updated} actuals saved", "period": period}


@router.put("/{project_id}/execution/progress/{prog_id}")
def update_progress(project_id: int, prog_id: int, body: dict,
                    db: Session = Depends(get_db),
                    current_user: User = Depends(get_current_user)):
    """Update a single plan or actual cell."""
    require_org_member(db, current_user.org_id, current_user.id, FINANCE_WRITE_ROLES)
    require_project_access(db, project_id, current_user.org_id)
    row = db.execute(text(
        "SELECT id FROM wf_progress WHERE id=:id AND project_id=:pid AND org_id=:oid"
    ), {"id": prog_id, "pid": project_id, "oid": current_user.org_id}).fetchone()
    if not row:
        raise HTTPException(404)
    sets = []
    vals: dict = {"id": prog_id}
    if "planned_qty" in body:
        sets.append("planned_qty=:plan"); vals["plan"] = body["planned_qty"]
    if "actual_qty" in body:
        sets.append("actual_qty=:act"); vals["act"] = body["actual_qty"]
        sets.append("entered_by=:uid"); vals["uid"] = current_user.id
        sets.append("entered_at=:now"); vals["now"] = datetime.utcnow().isoformat()
    if "notes" in body:
        sets.append("notes=:notes"); vals["notes"] = body["notes"]
    if sets:
        db.execute(text(f"UPDATE wf_progress SET {', '.join(sets)} WHERE id=:id"), vals)
        db.commit()
    return {"msg": "updated"}


# ─── Clear all data for a project ─────────────────────────────────────────────

@router.delete("/{project_id}/execution/clear")
def clear_execution_data(project_id: int, db: Session = Depends(get_db),
                          current_user: User = Depends(get_current_user)):
    """Delete all execution forecast data for this project (before re-importing)."""
    require_org_member(db, current_user.org_id, current_user.id, FINANCE_WRITE_ROLES)
    require_project_access(db, project_id, current_user.org_id)
    oid = current_user.org_id
    pid = project_id
    # Delete in FK order (progress → tasks → nodes + workstreams)
    db.execute(text(
        "DELETE FROM wf_progress WHERE project_id=:pid AND org_id=:oid"
    ), {"pid": pid, "oid": oid})
    db.execute(text(
        "DELETE FROM wf_tasks WHERE project_id=:pid AND org_id=:oid"
    ), {"pid": pid, "oid": oid})
    db.execute(text(
        "DELETE FROM wf_nodes WHERE project_id=:pid AND org_id=:oid"
    ), {"pid": pid, "oid": oid})
    db.execute(text(
        "DELETE FROM wf_workstreams WHERE project_id=:pid AND org_id=:oid"
    ), {"pid": pid, "oid": oid})
    db.commit()
    return {"msg": "All execution forecast data cleared for this project"}


# ─── Export back to Excel ─────────────────────────────────────────────────────

@router.get("/{project_id}/execution/export")
def export_excel(project_id: int, db: Session = Depends(get_db),
                 current_user: User = Depends(get_current_user)):
    """Download current plan + actuals as Excel (same column layout as import template)."""
    require_org_member(db, current_user.org_id, current_user.id, FINANCE_READ_ROLES)
    require_project_access(db, project_id, current_user.org_id)

    grid = get_grid(project_id, db=db, current_user=current_user)
    tasks = grid["tasks"]
    periods = [p["iso"] for p in grid["periods"]]

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Detailed Project Plan"

    hdr_fill = PatternFill("solid", fgColor="1F3864")
    hdr_font = Font(color="FFFFFF", bold=True, size=10)

    headers = ["Node", "SD", "Workstream", "Vendor", "Start Date", "Deadline",
               "Unit", "Total Scope", "Remaining"]
    for p in periods:
        lbl = _period_label(p)
        headers += [f"{lbl} Plan", f"{lbl} Actual"]

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")

    for ri, t in enumerate(tasks, 2):
        row_data = [
            t["node_code"], t["sd_code"], t["workstream"], t["vendor"],
            t["start_date"], t["deadline"], t["unit"],
            t["total_scope"], t["remaining"],
        ]
        for p in periods:
            pd = t["periods"].get(p, {"plan": 0, "actual": None})
            row_data += [pd["plan"] or 0, pd["actual"]]
        for ci, val in enumerate(row_data, 1):
            ws.cell(row=ri, column=ci, value=val)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=\"execution_forecast.xlsx\""},
    )
