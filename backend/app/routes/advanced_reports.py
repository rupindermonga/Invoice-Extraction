"""Advanced Reports: WIP Schedule Excel Export, Cash Requirement Forecast,
GL Journal Entry Export, Tax Mapping, Sub-tier Payment Visibility."""
import io, csv
from datetime import datetime, date, timedelta
from collections import defaultdict
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse, Response
from sqlalchemy.orm import Session

from ..database import get_db
from ..dependencies import get_current_user, get_current_org, FINANCE_READ_ROLES
from ..models import (
    Project, Invoice, Draw, ChangeOrder, CommittedCost, CostCategory,
    Payment, InvoiceAllocation, Organization,
)

router = APIRouter(prefix="/api/reports", tags=["advanced-reports"])


def _get_org(org_ctx, db):
    org, mem = org_ctx
    if mem.role not in FINANCE_READ_ROLES: raise HTTPException(403)
    return org


# ── WIP Schedule Excel Export ───────────────────────────────────────────────────

@router.get("/wip-schedule-excel")
def wip_schedule_excel(project_id: int = None, org_ctx=Depends(get_current_org),
                       db: Session = Depends(get_db), user=Depends(get_current_user)):
    """Export a GAAP-compliant WIP schedule in Excel format."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, numbers
    from openpyxl.utils import get_column_letter

    org = _get_org(org_ctx, db)
    projects = db.query(Project).filter(Project.org_id == org.id).all()
    if project_id:
        projects = [p for p in projects if p.id == project_id]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "WIP Schedule"

    # Header
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    headers = [
        "Project", "Contract Value", "Approved COs", "Revised Contract",
        "% Complete (Approved/Revised)", "Earned Revenue",
        "Total Invoiced", "Lender Approved", "Holdback",
        "Net Receivable", "Over-Billing (Liability)", "Under-Billing (Asset)",
        "Total Costs Incurred", "EAC", "Margin at Completion",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill; cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = 18

    row = 2
    for proj in projects:
        invoices = db.query(Invoice).filter(Invoice.project_id == proj.id, Invoice.status == "processed").all()
        approved_cos = db.query(ChangeOrder).filter(ChangeOrder.project_id == proj.id, ChangeOrder.status == "approved").all()
        co_total = sum(c.amount for c in approved_cos)
        revised = (proj.total_budget or 0) + co_total
        total_invoiced = sum(i.lender_submitted_amt or i.total_due or 0 for i in invoices)
        total_approved = sum(i.lender_approved_amt or 0 for i in invoices)
        holdback = sum((i.lender_approved_amt or 0) * (i.holdback_pct or 10) / 100 for i in invoices)
        pct = total_approved / revised if revised > 0 else 0
        earned = revised * pct
        over_billing = max(0, total_invoiced - earned)
        under_billing = max(0, earned - total_invoiced)
        committed = sum(c.contract_amount for c in db.query(CommittedCost).filter(CommittedCost.project_id == proj.id, CommittedCost.status == "active").all())
        eac = total_invoiced + committed
        margin = revised - eac

        values = [proj.name, proj.total_budget or 0, co_total, revised,
                  round(pct * 100, 1), round(earned, 2),
                  round(total_invoiced, 2), round(total_approved, 2), round(holdback, 2),
                  round(total_approved - holdback, 2), round(over_billing, 2), round(under_billing, 2),
                  round(total_invoiced, 2), round(eac, 2), round(margin, 2)]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            if col > 1: cell.number_format = '#,##0.00'
        row += 1

    ws.row_dimensions[1].height = 40
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return Response(
        content=output.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=WIP_Schedule.xlsx"},
    )


# ── Cash Requirement Forecast ────────────────────────────────────────────────────

@router.get("/cash-forecast")
def cash_forecast(project_id: int, weeks_ahead: int = 12,
                  org_ctx=Depends(get_current_org), db: Session = Depends(get_db)):
    """Weekly cash requirement forecast based on committed costs + draw schedule."""
    org = _get_org(org_ctx, db)
    proj = db.query(Project).filter(Project.id == project_id, Project.org_id == org.id).first()
    if not proj: raise HTTPException(404)

    today = date.today()
    weeks = []
    for i in range(weeks_ahead):
        week_start = today + timedelta(weeks=i)
        week_end = week_start + timedelta(days=6)
        weeks.append({"week": i + 1, "start": week_start.isoformat(), "end": week_end.isoformat(),
                       "invoices_due": 0, "holdback_released": 0, "net_required": 0})

    # Overlay unpaid invoices by due date
    unpaid = db.query(Invoice).filter(
        Invoice.project_id == project_id, Invoice.status == "processed",
        Invoice.payment_status != "paid", Invoice.due_date != None,
    ).all()
    for inv in unpaid:
        try:
            due = datetime.strptime(inv.due_date, "%Y-%m-%d").date()
            week_idx = (due - today).days // 7
            if 0 <= week_idx < weeks_ahead:
                amt = (inv.lender_approved_amt or inv.total_due or 0) - (inv.amount_paid or 0)
                weeks[week_idx]["invoices_due"] += amt
        except Exception:
            pass

    # Net required = invoices due
    for w in weeks:
        w["invoices_due"] = round(w["invoices_due"], 2)
        w["net_required"] = round(w["invoices_due"] - w["holdback_released"], 2)

    # Running cumulative
    cumulative = 0
    for w in weeks:
        cumulative += w["net_required"]
        w["cumulative"] = round(cumulative, 2)

    return {
        "project_name": proj.name, "forecast_weeks": weeks_ahead,
        "total_required": round(sum(w["net_required"] for w in weeks), 2),
        "weeks": weeks,
    }


# ── GL Journal Entry Export ───────────────────────────────────────────────────────

@router.get("/gl-export")
def gl_export(project_id: int = None, date_from: str = None, date_to: str = None,
              org_ctx=Depends(get_current_org), db: Session = Depends(get_db)):
    """Export general ledger journal entries for all processed invoices."""
    org = _get_org(org_ctx, db)
    q = db.query(Invoice).filter(Invoice.org_id == org.id, Invoice.status == "processed")
    if project_id: q = q.filter(Invoice.project_id == project_id)
    if date_from: q = q.filter(Invoice.invoice_date >= date_from)
    if date_to: q = q.filter(Invoice.invoice_date <= date_to)
    invoices = q.all()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Entry#","Date","Account Code","Account Name","Debit","Credit","Description","Reference","Project","Vendor"])
    entry_num = 1
    for inv in invoices:
        proj = db.query(Project).filter(Project.id == inv.project_id).first() if inv.project_id else None
        desc = f"{inv.vendor_name or ''} INV#{inv.invoice_number or inv.id}"
        date_str = inv.invoice_date or datetime.utcnow().strftime("%Y-%m-%d")
        subtotal = inv.subtotal or (inv.total_due or 0) - (inv.tax_total or 0)
        tax = inv.tax_total or 0
        # Debit: Construction in Progress (WIP asset)
        writer.writerow([entry_num, date_str, "1400", "Construction in Progress", round(subtotal, 2), "", desc, inv.invoice_number or "", proj.name if proj else "", inv.vendor_name or ""])
        if tax > 0:
            writer.writerow([entry_num, date_str, "1450", "Input Tax Credits (ITC)", round(tax, 2), "", f"GST/HST on {desc}", inv.invoice_number or "", proj.name if proj else "", inv.vendor_name or ""])
        # Credit: Accounts Payable
        writer.writerow([entry_num, date_str, "2000", "Accounts Payable", "", round(inv.total_due or 0, 2), desc, inv.invoice_number or "", proj.name if proj else "", inv.vendor_name or ""])
        entry_num += 1
        # If payment recorded
        for pmt in inv.payments:
            writer.writerow([entry_num, pmt.payment_date, "2000", "Accounts Payable", round(pmt.amount, 2), "", f"Payment: {desc}", pmt.reference or "", proj.name if proj else "", inv.vendor_name or ""])
            writer.writerow([entry_num, pmt.payment_date, "1010", "Cash / Bank", "", round(pmt.amount, 2), f"Payment: {desc}", pmt.reference or "", proj.name if proj else "", inv.vendor_name or ""])
            entry_num += 1

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]), media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=GL_Journal_Entries.csv"},
    )


# ── Tax Mapping Reference ─────────────────────────────────────────────────────────

@router.get("/tax-reference")
def tax_reference(province: str = None):
    """Canadian provincial tax rates and construction-specific rules."""
    PROVINCES = {
        "ON": {"name": "Ontario", "tax_type": "HST", "rate": 13.0, "federal_component": 5.0, "provincial_component": 8.0,
               "itc_eligible": True, "self_assess_required": False, "notes": "HST applies. 13% on most construction services."},
        "BC": {"name": "British Columbia", "tax_type": "GST+PST", "rate": 5.0, "pst_rate": 7.0, "total": 12.0,
               "itc_eligible": True, "pst_construction_note": "PST on materials, not labour. Construction services generally exempt from PST.",
               "notes": "GST 5% + PST 7%. Construction services: PST on materials only."},
        "AB": {"name": "Alberta", "tax_type": "GST", "rate": 5.0, "provincial_component": 0.0,
               "itc_eligible": True, "notes": "Alberta has no provincial sales tax. GST 5% only."},
        "QC": {"name": "Quebec", "tax_type": "GST+QST", "rate": 5.0, "qst_rate": 9.975, "total": 14.975,
               "itc_eligible": True, "qst_itc_eligible": True,
               "notes": "GST 5% + QST 9.975% = 14.975%. Both ITCs and QST refunds available for registered businesses."},
        "MB": {"name": "Manitoba", "tax_type": "GST+RST", "rate": 5.0, "rst_rate": 7.0, "total": 12.0,
               "itc_eligible": True, "notes": "GST 5% + RST (Retail Sales Tax) 7% on materials."},
        "SK": {"name": "Saskatchewan", "tax_type": "GST+PST", "rate": 5.0, "pst_rate": 6.0, "total": 11.0,
               "itc_eligible": True, "notes": "GST 5% + PST 6%. Construction contracts may be exempt; confirm with CRA."},
        "NS": {"name": "Nova Scotia", "tax_type": "HST", "rate": 15.0, "federal_component": 5.0, "provincial_component": 10.0,
               "itc_eligible": True, "notes": "HST 15%. Highest HST rate in Canada."},
        "NB": {"name": "New Brunswick", "tax_type": "HST", "rate": 15.0, "federal_component": 5.0, "provincial_component": 10.0,
               "itc_eligible": True, "notes": "HST 15%."},
        "NL": {"name": "Newfoundland & Labrador", "tax_type": "HST", "rate": 15.0, "federal_component": 5.0, "provincial_component": 10.0,
               "itc_eligible": True, "notes": "HST 15%."},
        "PE": {"name": "Prince Edward Island", "tax_type": "HST", "rate": 15.0, "federal_component": 5.0, "provincial_component": 10.0,
               "itc_eligible": True, "notes": "HST 15%."},
        "YT": {"name": "Yukon", "tax_type": "GST", "rate": 5.0, "notes": "GST only."},
        "NT": {"name": "Northwest Territories", "tax_type": "GST", "rate": 5.0, "notes": "GST only."},
        "NU": {"name": "Nunavut", "tax_type": "GST", "rate": 5.0, "notes": "GST only."},
    }
    if province:
        if province not in PROVINCES: raise HTTPException(404, f"Province {province} not found")
        return {province: PROVINCES[province]}
    return {"provinces": PROVINCES,
            "construction_notes": {
                "itc_general": "Businesses registered for GST/HST can claim Input Tax Credits (ITCs) on tax paid for business purposes.",
                "holdback_itc": "ITCs on holdback amounts can only be claimed when holdback is paid (cash basis).",
                "self_supply": "Owner-built residential real property may trigger self-supply rules — consult CRA.",
                "t5018": "Payments to sub-contractors >$500 require T5018 reporting to CRA.",
                "cra_guidance": "See CRA RC4058 for construction industry GST/HST guide.",
            }}


# ── Sub-Tier Payment Visibility ─────────────────────────────────────────────────

@router.get("/sub-tier-payments/{project_id}")
def sub_tier_payments(project_id: int, org_ctx=Depends(get_current_org), db: Session = Depends(get_db)):
    """View payment flow to sub-tier contractors — supply chain transparency."""
    org = _get_org(org_ctx, db)
    proj = db.query(Project).filter(Project.id == project_id, Project.org_id == org.id).first()
    if not proj: raise HTTPException(404)
    # Group invoices by vendor (first tier)
    invoices = db.query(Invoice).filter(Invoice.project_id == project_id, Invoice.status == "processed").all()
    vendors = defaultdict(lambda: {"total_billed": 0, "total_paid": 0, "unpaid": 0, "overdue": 0, "invoices": []})
    today = date.today().isoformat()
    for inv in invoices:
        vendor = inv.vendor_name or "Unknown"
        amt = inv.total_due or 0
        paid = inv.amount_paid or 0
        vendors[vendor]["total_billed"] += amt
        vendors[vendor]["total_paid"] += paid
        vendors[vendor]["unpaid"] += max(0, amt - paid)
        if inv.payment_status != "paid" and inv.due_date and inv.due_date < today:
            vendors[vendor]["overdue"] += max(0, amt - paid)
        vendors[vendor]["invoices"].append({
            "id": inv.id, "invoice_number": inv.invoice_number, "amount": amt,
            "paid": paid, "status": inv.payment_status, "due_date": inv.due_date,
        })
    rows = [{"vendor": k, "total_billed": round(v["total_billed"],2), "total_paid": round(v["total_paid"],2),
             "unpaid": round(v["unpaid"],2), "overdue": round(v["overdue"],2),
             "invoice_count": len(v["invoices"]),
             "payment_rate": round(v["total_paid"]/v["total_billed"]*100,1) if v["total_billed"] else 0}
            for k, v in vendors.items()]
    rows.sort(key=lambda x: -x["overdue"])
    return {
        "vendors": rows,
        "summary": {
            "total_vendors": len(rows),
            "total_billed": round(sum(r["total_billed"] for r in rows), 2),
            "total_paid": round(sum(r["total_paid"] for r in rows), 2),
            "total_unpaid": round(sum(r["unpaid"] for r in rows), 2),
            "total_overdue": round(sum(r["overdue"] for r in rows), 2),
        }
    }


# ── Forecast Accuracy Report ──────────────────────────────────────────────────────

@router.get("/forecast-accuracy/{project_id}")
def forecast_accuracy(project_id: int, org_ctx=Depends(get_current_org), db: Session = Depends(get_db)):
    """Compare prior EAC estimates to actuals — tracks forecast accuracy over time."""
    org = _get_org(org_ctx, db)
    proj = db.query(Project).filter(Project.id == project_id, Project.org_id == org.id).first()
    if not proj: raise HTTPException(404)
    invoices = db.query(Invoice).filter(Invoice.project_id == project_id, Invoice.status == "processed").all()
    approved_cos = db.query(ChangeOrder).filter(ChangeOrder.project_id == project_id, ChangeOrder.status == "approved").all()
    committed = db.query(CommittedCost).filter(CommittedCost.project_id == project_id, CommittedCost.status == "active").all()
    draws = db.query(Draw).filter(Draw.project_id == project_id).order_by(Draw.draw_number).all()
    revised = (proj.total_budget or 0) + sum(c.amount for c in approved_cos)
    actual_invoiced = sum(i.lender_submitted_amt or i.total_due or 0 for i in invoices)
    committed_total = sum(c.contract_amount for c in committed)
    eac = actual_invoiced + committed_total
    variance = revised - eac
    # Build draw-by-draw EAC history (simplified: assume each draw gives a data point)
    history = []
    cumulative = 0
    for d in draws:
        d_invoices = db.query(Invoice).filter(Invoice.draw_id == d.id).all()
        d_amt = sum(i.lender_submitted_amt or i.total_due or 0 for i in d_invoices)
        cumulative += d_amt
        pct = cumulative / revised * 100 if revised > 0 else 0
        history.append({
            "draw": d.draw_number, "date": d.submission_date, "status": d.status,
            "invoiced_to_date": round(cumulative, 2), "pct_complete": round(pct, 1),
            "implied_eac": round(cumulative / (pct / 100) if pct > 0 else revised, 2),
        })
    return {
        "project_name": proj.name, "original_budget": proj.total_budget or 0,
        "revised_contract": revised, "actual_invoiced": round(actual_invoiced, 2),
        "committed_remaining": round(committed_total, 2),
        "current_eac": round(eac, 2), "budget_variance": round(variance, 2),
        "variance_pct": round(variance / revised * 100, 1) if revised else 0,
        "draw_history": history,
    }
