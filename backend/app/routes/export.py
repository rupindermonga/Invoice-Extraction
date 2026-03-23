import json
import io
from fastapi import APIRouter, Depends, Header
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from ..database import get_db
from ..models import Invoice, ColumnConfig, User

router = APIRouter(prefix="/api/export", tags=["export"])


def _get_filtered_invoices(db, user_id, start_date, end_date, vendor, currency, status_filter):
    query = db.query(Invoice).filter(
        Invoice.user_id == user_id,
        Invoice.status == "processed"
    )
    if start_date:
        query = query.filter(Invoice.invoice_date >= start_date)
    if end_date:
        query = query.filter(Invoice.invoice_date <= end_date)
    if vendor:
        query = query.filter(Invoice.vendor_name.ilike(f"%{vendor}%"))
    if currency:
        query = query.filter(Invoice.currency == currency.upper())
    if status_filter:
        query = query.filter(Invoice.status == status_filter)
    return query.order_by(Invoice.invoice_date.desc()).all()


def _get_active_columns(db, user_id):
    """Return columns that are both active (in table) AND marked for export."""
    return (
        db.query(ColumnConfig)
        .filter(
            ColumnConfig.user_id == user_id,
            ColumnConfig.is_active == True,
            ColumnConfig.is_exportable == True,
        )
        .order_by(ColumnConfig.display_order)
        .all()
    )


_FORMULA_PREFIXES = ('=', '+', '-', '@', '\t', '\r')


def _safe_cell(val):
    """Prevent Excel formula injection by prefixing dangerous strings with a single quote."""
    if isinstance(val, str) and val and val[0] in _FORMULA_PREFIXES:
        return "'" + val
    return val


def _get_cell_value(invoice: Invoice, field_key: str):
    data = invoice.extracted_data or {}
    val = data.get(field_key)
    if val is None:
        # Fall back to indexed columns
        val = getattr(invoice, field_key, None)
    if isinstance(val, list):
        return json.dumps(val)
    return _safe_cell(val)


def _auth_from_header(authorization: str, db: Session) -> User:
    """Authenticate via Authorization: Bearer header only — no query-string tokens."""
    from ..dependencies import SECRET_KEY, ALGORITHM
    from jose import jwt
    if not authorization or not authorization.startswith("Bearer "):
        raise ValueError("missing Bearer token")
    raw_token = authorization[7:]
    payload = jwt.decode(raw_token, SECRET_KEY, algorithms=[ALGORITHM])
    user_id = int(payload.get("sub"))
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise ValueError("not found")
    if not user.is_active:
        raise ValueError("account disabled")
    return user


@router.get("/excel")
def export_excel(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    vendor: Optional[str] = None,
    currency: Optional[str] = None,
    authorization: Optional[str] = Header(None),
    db: Session = Depends(get_db),
):
    try:
        current_user = _auth_from_header(authorization, db)
    except Exception:
        from fastapi import HTTPException
        raise HTTPException(status_code=401, detail="Invalid or missing token")
    invoices = _get_filtered_invoices(db, current_user.id, start_date, end_date, vendor, currency, None)
    columns = _get_active_columns(db, current_user.id)

    wb = openpyxl.Workbook()

    # ── Summary sheet ─────────────────────────────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = "Invoices"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    alt_fill = PatternFill(start_color="F0F4F8", end_color="F0F4F8", fill_type="solid")

    # Write header row — exclude line_items column (goes to separate sheet)
    summary_cols = [c for c in columns if c.field_key != "line_items"]
    for col_idx, col in enumerate(summary_cols, start=1):
        cell = ws_summary.cell(row=1, column=col_idx, value=col.field_label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Write data rows
    for row_idx, invoice in enumerate(invoices, start=2):
        fill = alt_fill if row_idx % 2 == 0 else None
        for col_idx, col in enumerate(summary_cols, start=1):
            val = _get_cell_value(invoice, col.field_key)
            cell = ws_summary.cell(row=row_idx, column=col_idx, value=val)
            if fill:
                cell.fill = fill

    # Auto-width columns
    for col_idx in range(1, len(summary_cols) + 1):
        ws_summary.column_dimensions[get_column_letter(col_idx)].width = 22

    # ── Line items sheet ──────────────────────────────────────────────────────
    ws_lines = wb.create_sheet("Line Items")
    line_headers = ["Invoice ID", "Invoice #", "Vendor", "Currency",
                    "Line #", "SKU", "Description", "Quantity", "Unit",
                    "Unit Price", "Discount", "Tax Rate", "Line Total"]
    for col_idx, h in enumerate(line_headers, start=1):
        cell = ws_lines.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    line_row = 2
    for invoice in invoices:
        data = invoice.extracted_data or {}
        line_items = data.get("line_items", [])
        if isinstance(line_items, list):
            for item in line_items:
                if not isinstance(item, dict):
                    continue
                row_data = [
                    invoice.id,
                    invoice.invoice_number,
                    invoice.vendor_name,
                    invoice.currency,
                    item.get("line_no"),
                    item.get("sku"),
                    item.get("description"),
                    item.get("qty"),
                    item.get("unit"),
                    item.get("unit_price"),
                    item.get("discount_amount"),
                    item.get("tax_rate"),
                    item.get("line_total"),
                ]
                for col_idx, val in enumerate(row_data, start=1):
                    ws_lines.cell(row=line_row, column=col_idx, value=_safe_cell(val))
                line_row += 1

    for col_idx in range(1, len(line_headers) + 1):
        ws_lines.column_dimensions[get_column_letter(col_idx)].width = 22

    # Stream response
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"invoices_{start_date or 'all'}_{end_date or 'all'}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@router.get("/json")
def export_json(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    vendor: Optional[str] = None,
    currency: Optional[str] = None,
    authorization: Optional[str] = Header(None),
    db: Session = Depends(get_db),
):
    try:
        current_user = _auth_from_header(authorization, db)
    except Exception:
        from fastapi import HTTPException
        raise HTTPException(status_code=401, detail="Invalid or missing token")
    invoices = _get_filtered_invoices(db, current_user.id, start_date, end_date, vendor, currency, None)
    columns = _get_active_columns(db, current_user.id)

    result = []
    for inv in invoices:
        row = {
            "id": inv.id,
            "source": inv.source,
            "original_filename": inv.original_filename,
            "processed_at": inv.processed_at.isoformat() if inv.processed_at else None,
            "confidence_score": inv.confidence_score,
        }
        data = inv.extracted_data or {}
        for col in columns:
            val = data.get(col.field_key)
            if val is None:
                val = getattr(inv, col.field_key, None)
            row[col.field_key] = val
        result.append(row)

    buf = io.BytesIO(json.dumps(result, indent=2, default=str).encode("utf-8"))
    buf.seek(0)

    filename = f"invoices_{start_date or 'all'}_{end_date or 'all'}.json"
    return StreamingResponse(
        buf,
        media_type="application/json",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )
