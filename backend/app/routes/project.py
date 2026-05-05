"""Project finance routes: project CRUD, cost categories, sub-divisions, allocations, payments, dashboard, bookkeeping export."""
import os
import uuid
from fastapi import APIRouter, Depends, HTTPException, Header, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from typing import List, Optional
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime

from ..database import get_db
from ..models import (
    User, Project, SubDivision, CostCategory, CostSubCategory,
    SubDivisionBudget, Invoice, InvoiceAllocation, Payment,
    Draw, Claim, PayrollEntry, ChangeOrder, CommittedCost, Subcontractor,
    LenderToken, ProjectDocument, Milestone, LienWaiver,
)
from ..schemas import (
    ProjectCreate, ProjectUpdate, ProjectOut,
    SubDivisionOut, CostCategoryOut, CostCategoryCreate, CostCategoryUpdate,
    CostSubCategoryCreate, CostSubCategoryOut,
    SubDivisionBudgetSet, AllocationCreate, AllocationOut,
    PaymentCreate, PaymentOut,
    DrawCreate, DrawUpdate, DrawOut,
    ClaimCreate, ClaimUpdate, ClaimOut,
    InvoiceCostUpdate, PayrollEntryCreate, PayrollEntryUpdate, PayrollEntryOut,
)
from ..dependencies import get_current_user

router = APIRouter(prefix="/api/project", tags=["project"])


# ─── Project resolver dependency ─────────────────────────────────────────────
# Any endpoint that Depends on these automatically accepts ?project_id=N.
# If project_id is omitted, the user's first project is used (backwards compat).

def _get_proj(
    project_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> Optional[Project]:
    q = db.query(Project).filter(Project.user_id == current_user.id)
    if project_id:
        q = q.filter(Project.id == project_id)
    return q.order_by(Project.created_at).first()


def _req_proj(proj: Optional[Project] = Depends(_get_proj)) -> Project:
    if not proj:
        raise HTTPException(status_code=404, detail="Project not found")
    return proj


# ─── Project CRUD ────────────────────────────────────────────────────────────

@router.get("/list")
def list_projects(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    """List all projects for the current user."""
    return db.query(Project).filter(Project.user_id == current_user.id).order_by(Project.created_at).all()


@router.get("", response_model=Optional[ProjectOut])
def get_project(proj: Optional[Project] = Depends(_get_proj)):
    return proj


@router.post("", response_model=ProjectOut)
def create_project(body: ProjectCreate, project_type: Optional[str] = None, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    """Create a new project. Pass ?project_type= to auto-seed the category structure."""
    proj = Project(user_id=current_user.id, **body.model_dump())
    db.add(proj)
    db.commit()
    db.refresh(proj)
    if project_type:
        from ..seed_project import seed_project_template
        seed_project_template(db, proj.id, project_type)
        db.refresh(proj)
    return proj


@router.put("", response_model=ProjectOut)
def update_project(body: ProjectUpdate, proj: Project = Depends(_req_proj), db: Session = Depends(get_db)):
    _ALLOWED = {"name", "code", "client", "address", "start_date", "end_date", "total_budget", "lender_budget", "currency"}
    for field, value in body.model_dump(exclude_unset=True).items():
        if field in _ALLOWED:
            setattr(proj, field, value)
    db.commit()
    db.refresh(proj)
    return proj


@router.post("/{project_id}/seed-template")
def apply_template(project_id: int, body: dict, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    """Apply a category template to an existing project.
    Only seeds if the project has no categories yet (safe to call on empty projects)."""
    proj = db.query(Project).filter(Project.id == project_id, Project.user_id == current_user.id).first()
    if not proj:
        raise HTTPException(status_code=404, detail="Project not found")
    project_type = body.get("project_type", "custom")
    from ..seed_project import seed_project_template, _TEMPLATES
    if project_type not in _TEMPLATES:
        raise HTTPException(status_code=400, detail=f"Unknown project_type. Choose: {', '.join(_TEMPLATES.keys())}")
    seed_project_template(db, proj.id, project_type)
    return {"message": f"Template '{project_type}' applied", "project_id": proj.id}


@router.delete("/{project_id}")
def delete_project(project_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    proj = db.query(Project).filter(Project.id == project_id, Project.user_id == current_user.id).first()
    if not proj:
        raise HTTPException(status_code=404, detail="Project not found")
    db.delete(proj)
    db.commit()
    return {"message": "Project deleted"}


# ─── Sub-Divisions ───────────────────────────────────────────────────────────

@router.get("/subdivisions", response_model=List[SubDivisionOut])
def list_subdivisions(proj: Optional[Project] = Depends(_get_proj), db: Session = Depends(get_db)):
    if not proj:
        return []
    return db.query(SubDivision).filter(SubDivision.project_id == proj.id).order_by(SubDivision.display_order).all()


@router.post("/subdivisions", response_model=SubDivisionOut)
def create_subdivision(name: str, description: str = None, proj: Project = Depends(_req_proj), db: Session = Depends(get_db)):
    if not proj:
        raise HTTPException(status_code=404, detail="Create a project first")
    sd = SubDivision(project_id=proj.id, name=name, description=description)
    db.add(sd)
    db.commit()
    db.refresh(sd)
    return sd


# ─── Cost Categories ─────────────────────────────────────────────────────────

@router.get("/categories", response_model=List[CostCategoryOut])
def list_cost_categories(proj: Optional[Project] = Depends(_get_proj), db: Session = Depends(get_db)):
    if not proj:
        return []
    return (
        db.query(CostCategory)
        .filter(CostCategory.project_id == proj.id)
        .order_by(CostCategory.display_order)
        .all()
    )


@router.post("/categories", response_model=CostCategoryOut)
def create_cost_category(body: CostCategoryCreate, proj: Project = Depends(_req_proj), db: Session = Depends(get_db)):
    if body.budget < 0:
        raise HTTPException(status_code=400, detail="Budget must be non-negative")
    existing = db.query(CostCategory).filter(CostCategory.project_id == proj.id, CostCategory.name == body.name).first()
    if existing:
        raise HTTPException(status_code=400, detail="Category name already exists in this project")
    cat = CostCategory(project_id=proj.id, **body.model_dump())
    db.add(cat)
    db.commit()
    db.refresh(cat)
    return cat


@router.put("/categories/{cat_id}", response_model=CostCategoryOut)
def update_cost_category(cat_id: int, body: CostCategoryUpdate, proj: Project = Depends(_req_proj), db: Session = Depends(get_db)):
    cat = db.query(CostCategory).filter(CostCategory.id == cat_id, CostCategory.project_id == proj.id).first()
    if not cat:
        raise HTTPException(status_code=404, detail="Cost category not found")
    updates = body.model_dump(exclude_unset=True)
    if "budget" in updates and updates["budget"] is not None and updates["budget"] < 0:
        raise HTTPException(status_code=400, detail="Budget must be non-negative")
    _ALLOWED = {"name", "budget", "lender_budget"}
    for field, value in updates.items():
        if field in _ALLOWED:
            setattr(cat, field, value)
    db.commit()
    db.refresh(cat)
    return cat


@router.delete("/categories/{cat_id}")
def delete_cost_category(cat_id: int, proj: Project = Depends(_req_proj), db: Session = Depends(get_db)):
    cat = db.query(CostCategory).filter(CostCategory.id == cat_id, CostCategory.project_id == proj.id).first()
    if not cat:
        raise HTTPException(status_code=404, detail="Cost category not found")
    db.delete(cat)
    db.commit()
    return {"message": "Deleted"}


# ─── Cost Sub-Categories ─────────────────────────────────────────────────────

@router.post("/categories/{cat_id}/subcategories", response_model=CostSubCategoryOut)
def create_cost_subcategory(cat_id: int, body: CostSubCategoryCreate, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    cat = (db.query(CostCategory).join(Project, Project.id == CostCategory.project_id)
            .filter(CostCategory.id == cat_id, Project.user_id == current_user.id).first())
    if not cat:
        raise HTTPException(status_code=404, detail="Cost category not found")
    sc = CostSubCategory(category_id=cat_id, name=body.name, description=body.description, budget=body.budget)
    db.add(sc)
    db.commit()
    db.refresh(sc)
    return sc


@router.delete("/subcategories/{sc_id}")
def delete_cost_subcategory(sc_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    sc = (db.query(CostSubCategory)
          .join(CostCategory, CostCategory.id == CostSubCategory.category_id)
          .join(Project, Project.id == CostCategory.project_id)
          .filter(CostSubCategory.id == sc_id, Project.user_id == current_user.id).first())
    if not sc:
        raise HTTPException(status_code=404)
    db.delete(sc)
    db.commit()
    return {"message": "Deleted"}


# ─── Sub-Division Budgets (for Fiber Build etc.) ─────────────────────────────

@router.put("/categories/{cat_id}/subdivision-budgets")
def set_subdivision_budgets(cat_id: int, budgets: List[SubDivisionBudgetSet], db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    cat = (db.query(CostCategory).join(Project, Project.id == CostCategory.project_id)
            .filter(CostCategory.id == cat_id, Project.user_id == current_user.id).first())
    if not cat:
        raise HTTPException(status_code=404)
    proj = db.query(Project).filter(Project.id == cat.project_id).first()
    # Upsert budgets — verify each subdivision belongs to this project
    for b in budgets:
        sd = db.query(SubDivision).filter(SubDivision.id == b.subdivision_id, SubDivision.project_id == proj.id).first()
        if not sd:
            raise HTTPException(status_code=404, detail=f"Sub-division {b.subdivision_id} not found in your project")
        if b.budget < 0:
            raise HTTPException(status_code=400, detail="Budget must be non-negative")
        existing = db.query(SubDivisionBudget).filter(
            SubDivisionBudget.category_id == cat_id,
            SubDivisionBudget.subdivision_id == b.subdivision_id
        ).first()
        if existing:
            existing.budget = b.budget
        else:
            db.add(SubDivisionBudget(category_id=cat_id, subdivision_id=b.subdivision_id, budget=b.budget))
    db.commit()
    return {"message": "Budgets updated"}


@router.get("/categories/{cat_id}/subdivision-budgets")
def get_subdivision_budgets(cat_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    cat = (db.query(CostCategory).join(Project, Project.id == CostCategory.project_id)
            .filter(CostCategory.id == cat_id, Project.user_id == current_user.id).first())
    if not cat:
        raise HTTPException(status_code=404, detail="Cost category not found")
    rows = db.query(SubDivisionBudget).filter(SubDivisionBudget.category_id == cat_id).all()
    return [{"subdivision_id": r.subdivision_id, "budget": r.budget} for r in rows]


# ─── Invoice Allocations ─────────────────────────────────────────────────────

@router.get("/allocations/{invoice_id}", response_model=List[AllocationOut])
def get_allocations(invoice_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    inv = db.query(Invoice).filter(Invoice.id == invoice_id, Invoice.user_id == current_user.id).first()
    if not inv:
        raise HTTPException(status_code=404)
    allocs = db.query(InvoiceAllocation).filter(InvoiceAllocation.invoice_id == invoice_id).all()
    result = []
    for a in allocs:
        out = AllocationOut(
            id=a.id, invoice_id=a.invoice_id, category_id=a.category_id,
            sub_category_id=a.sub_category_id, subdivision_id=a.subdivision_id,
            percentage=a.percentage, amount=a.amount,
            category_name=a.category.name if a.category else None,
            sub_category_name=a.sub_category.name if a.sub_category else None,
            subdivision_name=a.subdivision.name if a.subdivision else None,
        )
        result.append(out)
    return result


@router.put("/allocations/{invoice_id}")
def set_allocations(invoice_id: int, allocations: List[AllocationCreate], db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    """Replace all allocations for an invoice."""
    inv = db.query(Invoice).filter(Invoice.id == invoice_id, Invoice.user_id == current_user.id).first()
    if not inv:
        raise HTTPException(status_code=404)
    proj = (db.query(Project).filter(Project.user_id == current_user.id)
            .order_by(Project.created_at).first())

    # Validate individual percentages
    for a in allocations:
        if a.percentage <= 0 or a.percentage > 100:
            raise HTTPException(status_code=400, detail="Each allocation percentage must be between 0 and 100")

    # Validate total percentage
    total_pct = sum(a.percentage for a in allocations)
    if allocations and abs(total_pct - 100.0) > 0.01:
        raise HTTPException(status_code=400, detail=f"Allocation percentages must total 100% (got {total_pct}%)")

    # Validate all referenced cost structures belong to user's project
    for a in allocations:
        cat = db.query(CostCategory).filter(CostCategory.id == a.category_id, CostCategory.project_id == proj.id).first()
        if not cat:
            raise HTTPException(status_code=404, detail=f"Cost category {a.category_id} not found in your project")
        if a.sub_category_id:
            sc = db.query(CostSubCategory).filter(CostSubCategory.id == a.sub_category_id, CostSubCategory.category_id == a.category_id).first()
            if not sc:
                raise HTTPException(status_code=404, detail="Sub-category not found in this category")
        if a.subdivision_id:
            sd = db.query(SubDivision).filter(SubDivision.id == a.subdivision_id, SubDivision.project_id == proj.id).first()
            if not sd:
                raise HTTPException(status_code=404, detail="Sub-division not found in your project")

    # Clear old allocations
    db.query(InvoiceAllocation).filter(InvoiceAllocation.invoice_id == invoice_id).delete()

    # Create new
    invoice_total = inv.total_due or 0.0
    for a in allocations:
        amount = round(invoice_total * a.percentage / 100.0, 2)
        db.add(InvoiceAllocation(
            invoice_id=invoice_id,
            category_id=a.category_id,
            sub_category_id=a.sub_category_id,
            subdivision_id=a.subdivision_id,
            percentage=a.percentage,
            amount=amount,
        ))
    db.commit()
    return {"message": "Allocations saved", "count": len(allocations)}


# ─── Payments ─────────────────────────────────────────────────────────────────

@router.get("/payments/{invoice_id}", response_model=List[PaymentOut])
def list_payments(invoice_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    inv = db.query(Invoice).filter(Invoice.id == invoice_id, Invoice.user_id == current_user.id).first()
    if not inv:
        raise HTTPException(status_code=404)
    return db.query(Payment).filter(Payment.invoice_id == invoice_id).order_by(Payment.payment_date).all()


@router.post("/payments", response_model=PaymentOut)
def create_payment(body: PaymentCreate, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    inv = db.query(Invoice).filter(Invoice.id == body.invoice_id, Invoice.user_id == current_user.id).first()
    if not inv:
        raise HTTPException(status_code=404, detail="Invoice not found")
    if body.amount <= 0:
        raise HTTPException(status_code=400, detail="Payment amount must be positive")
    balance = (inv.total_due or 0.0) - (inv.amount_paid or 0.0)
    if balance <= 0:
        raise HTTPException(status_code=400, detail="Invoice is already fully paid")
    if body.amount > balance + 0.01:
        raise HTTPException(status_code=400, detail=f"Payment ${body.amount:.2f} exceeds outstanding balance ${balance:.2f}")

    pmt = Payment(
        invoice_id=body.invoice_id,
        amount=body.amount,
        payment_date=body.payment_date,
        method=body.method,
        reference=body.reference,
        notes=body.notes,
    )
    db.add(pmt)

    # Update invoice payment status
    inv.amount_paid = (inv.amount_paid or 0.0) + body.amount
    total = inv.total_due or 0.0
    if total > 0 and inv.amount_paid >= total - 0.01:
        inv.payment_status = "paid"
    elif inv.amount_paid > 0:
        inv.payment_status = "partially_paid"

    db.commit()
    db.refresh(pmt)
    return pmt


@router.delete("/payments/{payment_id}")
def delete_payment(payment_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    # Join through Invoice to verify ownership in a single query
    pmt = db.query(Payment).join(Invoice).filter(
        Payment.id == payment_id,
        Invoice.user_id == current_user.id,
    ).first()
    if not pmt:
        raise HTTPException(status_code=404, detail="Payment not found")
    inv = db.query(Invoice).filter(Invoice.id == pmt.invoice_id).first()
    if not inv:
        raise HTTPException(status_code=404)

    inv.amount_paid = max(0.0, (inv.amount_paid or 0.0) - pmt.amount)
    if inv.amount_paid <= 0.01:
        inv.payment_status = "unpaid"
        inv.amount_paid = 0.0
    else:
        inv.payment_status = "partially_paid"

    db.delete(pmt)
    db.commit()
    return {"message": "Payment deleted"}


# ─── Draws ───────────────────────────────────────────────────────────────────

def _draw_out(draw, db):
    invs = db.query(Invoice).filter(Invoice.draw_id == draw.id).all()
    total_orig = sum(i.total_due or 0 for i in invs)
    return DrawOut(
        id=draw.id, draw_number=draw.draw_number, fx_rate=draw.fx_rate,
        submission_date=draw.submission_date, status=draw.status,
        notes=draw.notes, created_at=draw.created_at,
        invoice_count=len(invs), total_original=round(total_orig, 2),
        total_cad=round(total_orig * draw.fx_rate, 2),
    )

@router.get("/draws")
def list_draws(proj: Optional[Project] = Depends(_get_proj), db: Session = Depends(get_db)):
    if not proj:
        return []
    draws = db.query(Draw).filter(Draw.project_id == proj.id).order_by(Draw.draw_number).all()
    return [_draw_out(d, db) for d in draws]

@router.post("/draws")
def create_draw(body: DrawCreate, proj: Project = Depends(_req_proj), db: Session = Depends(get_db)):
    existing = db.query(Draw).filter(Draw.project_id == proj.id, Draw.draw_number == body.draw_number).first()
    if existing:
        raise HTTPException(status_code=400, detail=f"Draw {body.draw_number} already exists")
    draw = Draw(project_id=proj.id, **body.model_dump())
    db.add(draw)
    db.commit()
    db.refresh(draw)
    return _draw_out(draw, db)

@router.put("/draws/{draw_id}")
def update_draw(draw_id: int, body: DrawUpdate, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    draw = (db.query(Draw).join(Project, Project.id == Draw.project_id)
            .filter(Draw.id == draw_id, Project.user_id == current_user.id).first())
    if not draw:
        raise HTTPException(status_code=404, detail="Draw not found")
    for field, value in body.model_dump(exclude_unset=True).items():
        if field in {"fx_rate", "submission_date", "status", "notes"}:
            setattr(draw, field, value)
    db.commit()
    db.refresh(draw)
    return _draw_out(draw, db)

@router.delete("/draws/{draw_id}")
def delete_draw(draw_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    draw = (db.query(Draw).join(Project, Project.id == Draw.project_id)
            .filter(Draw.id == draw_id, Project.user_id == current_user.id).first())
    if not draw:
        raise HTTPException(status_code=404)
    db.query(Invoice).filter(Invoice.draw_id == draw_id).update({"draw_id": None})
    db.delete(draw)
    db.commit()
    return {"message": "Draw deleted"}

@router.put("/draws/{draw_id}/invoices")
def assign_invoices_to_draw(draw_id: int, invoice_ids: List[int], db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    draw = (db.query(Draw).join(Project, Project.id == Draw.project_id)
            .filter(Draw.id == draw_id, Project.user_id == current_user.id).first())
    if not draw:
        raise HTTPException(status_code=404, detail="Draw not found")
    # Unlink all currently linked
    db.query(Invoice).filter(Invoice.draw_id == draw_id).update({"draw_id": None})
    # Link new set
    for iid in invoice_ids:
        inv = db.query(Invoice).filter(Invoice.id == iid, Invoice.user_id == current_user.id).first()
        if inv:
            inv.draw_id = draw_id
    db.commit()
    return _draw_out(draw, db)

@router.get("/draws/{draw_id}/invoices")
def get_draw_invoices(draw_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    draw = (db.query(Draw).join(Project, Project.id == Draw.project_id)
            .filter(Draw.id == draw_id, Project.user_id == current_user.id).first())
    if not draw:
        raise HTTPException(status_code=404)
    invs = db.query(Invoice).filter(Invoice.draw_id == draw_id, Invoice.user_id == current_user.id).all()
    return [{
        "id": i.id, "invoice_number": i.invoice_number, "vendor_name": i.vendor_name,
        "currency": i.currency or "CAD", "total_due": i.total_due or 0,
        "cad_amount": round((i.total_due or 0) * draw.fx_rate, 2) if (i.currency or "CAD").upper() != "CAD" else (i.total_due or 0),
        "original_filename": i.original_filename, "invoice_date": i.invoice_date,
        "billed_to": i.billed_to, "billing_type": i.billing_type, "vendor_on_record": i.vendor_on_record,
    } for i in invs]


# ─── Bulk Approve ────────────────────────────────────────────────────────────

@router.post("/draws/{draw_id}/approve-all")
def bulk_approve_draw(draw_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    """Approve all invoices in a draw: set lender_approved_amt = lender_submitted_amt, status = approved."""
    draw = (db.query(Draw).join(Project, Project.id == Draw.project_id)
            .filter(Draw.id == draw_id, Project.user_id == current_user.id).first())
    if not draw:
        raise HTTPException(status_code=404, detail="Draw not found")
    invs = db.query(Invoice).filter(Invoice.draw_id == draw_id, Invoice.user_id == current_user.id).all()
    count = 0
    for inv in invs:
        # Auto-calc submitted if not set
        st = inv.subtotal or inv.total_due or 0
        if inv.lender_submitted_amt is None:
            inv.lender_margin_amt = round(st * (inv.lender_margin_pct or 0) / 100, 2)
            inv.lender_submitted_amt = round(st + (inv.lender_margin_amt or 0) + (_calc_lender_tax(inv) if hasattr(inv, 'billing_type') else 0), 2)
        inv.lender_approved_amt = inv.lender_submitted_amt
        inv.lender_status = "approved"
        count += 1
    db.commit()
    return {"message": f"Approved {count} invoices", "count": count}


@router.post("/claims/{claim_id}/approve-all")
def bulk_approve_claim(claim_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    """Approve all invoices in a claim: set govt_approved_amt = govt_submitted_amt, status = approved."""
    claim = (db.query(Claim).join(Project, Project.id == Claim.project_id)
             .filter(Claim.id == claim_id, Project.user_id == current_user.id).first())
    if not claim:
        raise HTTPException(status_code=404, detail="Claim not found")
    invs = db.query(Invoice).filter(_claim_fk(claim) == claim_id, Invoice.user_id == current_user.id).all()
    count = 0
    for inv in invs:
        st = inv.subtotal or inv.total_due or 0
        if inv.govt_submitted_amt is None:
            inv.govt_margin_amt = round(st * (inv.govt_margin_pct or 0) / 100, 2)
            inv.govt_submitted_amt = round(st + (inv.govt_margin_amt or 0), 2)
        inv.govt_approved_amt = inv.govt_submitted_amt
        inv.govt_status = "approved"
        count += 1
    db.commit()
    return {"message": f"Approved {count} invoices", "count": count}


# ─── Claims ──────────────────────────────────────────────────────────────────

def _claim_fk(claim):
    """Return the Invoice FK column for this claim's type."""
    return Invoice.provincial_claim_id if claim.claim_type == "provincial" else Invoice.federal_claim_id

def _claim_fk_name(claim):
    """Return the Invoice FK column name string for this claim's type."""
    return "provincial_claim_id" if claim.claim_type == "provincial" else "federal_claim_id"

def _claim_out(claim, db):
    invs = db.query(Invoice).filter(_claim_fk(claim) == claim.id).all()
    total_orig = sum(i.total_due or 0 for i in invs)
    return ClaimOut(
        id=claim.id, claim_number=claim.claim_number, claim_type=claim.claim_type,
        fx_rate=claim.fx_rate, submission_date=claim.submission_date,
        status=claim.status, notes=claim.notes, created_at=claim.created_at,
        invoice_count=len(invs), total_original=round(total_orig, 2),
        total_cad=round(total_orig * claim.fx_rate, 2),
    )

@router.get("/claims")
def list_claims(claim_type: Optional[str] = None, proj: Optional[Project] = Depends(_get_proj), db: Session = Depends(get_db)):
    if not proj:
        return []
    q = db.query(Claim).filter(Claim.project_id == proj.id)
    if claim_type:
        q = q.filter(Claim.claim_type == claim_type)
    return [_claim_out(c, db) for c in q.order_by(Claim.claim_number).all()]

@router.post("/claims")
def create_claim(body: ClaimCreate, proj: Project = Depends(_req_proj), db: Session = Depends(get_db)):
    if body.claim_type not in ("provincial", "federal"):
        raise HTTPException(status_code=400, detail="claim_type must be 'provincial' or 'federal'")
    existing = db.query(Claim).filter(Claim.project_id == proj.id, Claim.claim_number == body.claim_number, Claim.claim_type == body.claim_type).first()
    if existing:
        raise HTTPException(status_code=400, detail=f"{body.claim_type.title()} Claim {body.claim_number} already exists")
    claim = Claim(project_id=proj.id, **body.model_dump())
    db.add(claim)
    db.commit()
    db.refresh(claim)
    return _claim_out(claim, db)

@router.put("/claims/{claim_id}")
def update_claim(claim_id: int, body: ClaimUpdate, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    claim = (db.query(Claim).join(Project, Project.id == Claim.project_id)
             .filter(Claim.id == claim_id, Project.user_id == current_user.id).first())
    if not claim:
        raise HTTPException(status_code=404, detail="Claim not found")
    for field, value in body.model_dump(exclude_unset=True).items():
        if field in {"fx_rate", "submission_date", "status", "notes"}:
            setattr(claim, field, value)
    db.commit()
    db.refresh(claim)
    return _claim_out(claim, db)

@router.delete("/claims/{claim_id}")
def delete_claim(claim_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    claim = (db.query(Claim).join(Project, Project.id == Claim.project_id)
             .filter(Claim.id == claim_id, Project.user_id == current_user.id).first())
    if not claim:
        raise HTTPException(status_code=404)
    fk_name = _claim_fk_name(claim)
    db.query(Invoice).filter(_claim_fk(claim) == claim_id).update({fk_name: None})
    db.delete(claim)
    db.commit()
    return {"message": "Claim deleted"}

@router.put("/claims/{claim_id}/invoices")
def assign_invoices_to_claim(claim_id: int, invoice_ids: List[int], db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    claim = (db.query(Claim).join(Project, Project.id == Claim.project_id)
             .filter(Claim.id == claim_id, Project.user_id == current_user.id).first())
    if not claim:
        raise HTTPException(status_code=404, detail="Claim not found")
    fk_name = _claim_fk_name(claim)
    db.query(Invoice).filter(_claim_fk(claim) == claim_id).update({fk_name: None})
    for iid in invoice_ids:
        inv = db.query(Invoice).filter(Invoice.id == iid, Invoice.user_id == current_user.id).first()
        if inv:
            setattr(inv, fk_name, claim_id)
    db.commit()
    return _claim_out(claim, db)

@router.put("/claims/{claim_id}/copy-from-draw/{draw_id}")
def copy_draw_to_claim(claim_id: int, draw_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    """Copy all invoices from a draw to a claim (for the 99% overlap case)."""
    draw = (db.query(Draw).join(Project, Project.id == Draw.project_id)
            .filter(Draw.id == draw_id, Project.user_id == current_user.id).first())
    claim = (db.query(Claim).join(Project, Project.id == Claim.project_id)
             .filter(Claim.id == claim_id, Project.user_id == current_user.id).first())
    if not draw or not claim:
        raise HTTPException(status_code=404)
    fk_name = _claim_fk_name(claim)
    draw_invs = db.query(Invoice).filter(Invoice.draw_id == draw_id, Invoice.user_id == current_user.id).all()
    for inv in draw_invs:
        setattr(inv, fk_name, claim_id)
    db.commit()
    return _claim_out(claim, db)

@router.get("/claims/{claim_id}/invoices")
def get_claim_invoices(claim_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    claim = (db.query(Claim).join(Project, Project.id == Claim.project_id)
             .filter(Claim.id == claim_id, Project.user_id == current_user.id).first())
    if not claim:
        raise HTTPException(status_code=404)
    invs = db.query(Invoice).filter(_claim_fk(claim) == claim_id, Invoice.user_id == current_user.id).all()
    return [{
        "id": i.id, "invoice_number": i.invoice_number, "vendor_name": i.vendor_name,
        "currency": i.currency or "CAD", "total_due": i.total_due or 0,
        "cad_amount": round((i.total_due or 0) * claim.fx_rate, 2) if (i.currency or "CAD").upper() != "CAD" else (i.total_due or 0),
        "original_filename": i.original_filename, "invoice_date": i.invoice_date,
        "billed_to": i.billed_to, "billing_type": i.billing_type, "vendor_on_record": i.vendor_on_record,
    } for i in invs]


# ─── FX Rate (Bank of Canada) ───────────────────────────────────────────────

@router.get("/fx-rate")
def get_fx_rate(date: Optional[str] = None):
    """Fetch USD→CAD rate from Bank of Canada for a given date. Falls back to 1.0 on error."""
    import urllib.request, json as _json
    target_date = date or datetime.utcnow().strftime("%Y-%m-%d")
    try:
        url = f"https://www.bankofcanada.ca/valet/observations/FXUSDCAD/json?start_date={target_date}&end_date={target_date}"
        with urllib.request.urlopen(url, timeout=5) as resp:
            data = _json.loads(resp.read())
        obs = data.get("observations", [])
        if obs:
            return {"date": target_date, "rate": float(obs[-1]["FXUSDCAD"]["v"]), "source": "bank_of_canada"}
        # If no data for that date (weekend/holiday), try last 5 days
        from datetime import timedelta
        d = datetime.strptime(target_date, "%Y-%m-%d")
        start = (d - timedelta(days=5)).strftime("%Y-%m-%d")
        url2 = f"https://www.bankofcanada.ca/valet/observations/FXUSDCAD/json?start_date={start}&end_date={target_date}"
        with urllib.request.urlopen(url2, timeout=5) as resp:
            data = _json.loads(resp.read())
        obs = data.get("observations", [])
        if obs:
            last = obs[-1]
            return {"date": last["d"], "rate": float(last["FXUSDCAD"]["v"]), "source": "bank_of_canada"}
    except Exception:
        pass
    return {"date": target_date, "rate": 1.0, "source": "fallback"}


# ─── Change Orders ───────────────────────────────────────────────────────────

@router.get("/change-orders")
def list_change_orders(proj: Optional[Project] = Depends(_get_proj), db: Session = Depends(get_db)):
    if not proj:
        return []
    return [
        {
            "id": co.id, "co_number": co.co_number, "description": co.description,
            "amount": co.amount, "status": co.status, "issued_by": co.issued_by,
            "date": co.date, "notes": co.notes, "created_at": str(co.created_at),
            "category_id": co.category_id,
            "category_name": co.category.name if co.category else None,
        }
        for co in db.query(ChangeOrder)
        .filter(ChangeOrder.project_id == proj.id)
        .order_by(ChangeOrder.date.desc(), ChangeOrder.id.desc())
        .all()
    ]


@router.post("/change-orders")
def create_change_order(body: dict, proj: Project = Depends(_req_proj), db: Session = Depends(get_db)):
    if not body.get("co_number"):
        raise HTTPException(status_code=400, detail="co_number is required")
    if not body.get("description"):
        raise HTTPException(status_code=400, detail="description is required")
    if body.get("amount") is None:
        raise HTTPException(status_code=400, detail="amount is required")
    cat_id = body.get("category_id")
    if cat_id:
        cat = db.query(CostCategory).filter(CostCategory.id == cat_id, CostCategory.project_id == proj.id).first()
        if not cat:
            raise HTTPException(status_code=404, detail="Cost category not found in this project")
    co = ChangeOrder(
        project_id=proj.id,
        category_id=cat_id,
        co_number=body["co_number"],
        description=body["description"],
        amount=float(body["amount"]),
        status=body.get("status", "pending"),
        issued_by=body.get("issued_by"),
        date=body.get("date"),
        notes=body.get("notes"),
    )
    db.add(co)
    db.commit()
    db.refresh(co)
    return {"id": co.id, "co_number": co.co_number, "amount": co.amount, "status": co.status}


@router.put("/change-orders/{co_id}")
def update_change_order(co_id: int, body: dict, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    co = (db.query(ChangeOrder).join(Project, Project.id == ChangeOrder.project_id)
          .filter(ChangeOrder.id == co_id, Project.user_id == current_user.id).first())
    if not co:
        raise HTTPException(status_code=404, detail="Change order not found")
    for field in ("co_number", "description", "amount", "status", "issued_by", "date", "notes", "category_id"):
        if field in body:
            setattr(co, field, body[field])
    db.commit()
    db.refresh(co)
    return {"id": co.id, "co_number": co.co_number, "amount": co.amount, "status": co.status}


@router.delete("/change-orders/{co_id}")
def delete_change_order(co_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    co = (db.query(ChangeOrder).join(Project, Project.id == ChangeOrder.project_id)
          .filter(ChangeOrder.id == co_id, Project.user_id == current_user.id).first())
    if not co:
        raise HTTPException(status_code=404, detail="Change order not found")
    db.delete(co)
    db.commit()
    return {"message": "Deleted"}


# ─── Milestones ──────────────────────────────────────────────────────────────

@router.get("/milestones")
def list_milestones(proj: Optional[Project] = Depends(_get_proj), db: Session = Depends(get_db)):
    if not proj: return []
    return [
        {"id": m.id, "name": m.name, "description": m.description, "target_date": m.target_date,
         "actual_date": m.actual_date, "pct_complete": m.pct_complete, "status": m.status,
         "display_order": m.display_order}
        for m in db.query(Milestone).filter(Milestone.project_id == proj.id).order_by(Milestone.display_order).all()
    ]


@router.post("/milestones")
def create_milestone(body: dict, proj: Project = Depends(_req_proj), db: Session = Depends(get_db)):
    if not body.get("name"): raise HTTPException(status_code=400, detail="name required")
    m = Milestone(project_id=proj.id, name=body["name"], description=body.get("description"),
                  target_date=body.get("target_date"), actual_date=body.get("actual_date"),
                  pct_complete=float(body.get("pct_complete", 0)), status=body.get("status", "pending"),
                  display_order=int(body.get("display_order", 100)))
    db.add(m); db.commit(); db.refresh(m)
    return {"id": m.id, "name": m.name, "status": m.status}


@router.put("/milestones/{ms_id}")
def update_milestone(ms_id: int, body: dict, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    m = (db.query(Milestone).join(Project, Project.id == Milestone.project_id)
         .filter(Milestone.id == ms_id, Project.user_id == current_user.id).first())
    if not m: raise HTTPException(status_code=404)
    for f in ("name","description","target_date","actual_date","pct_complete","status","display_order"):
        if f in body: setattr(m, f, body[f])
    db.commit(); db.refresh(m)
    return {"id": m.id, "pct_complete": m.pct_complete, "status": m.status}


@router.delete("/milestones/{ms_id}")
def delete_milestone(ms_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    m = (db.query(Milestone).join(Project, Project.id == Milestone.project_id)
         .filter(Milestone.id == ms_id, Project.user_id == current_user.id).first())
    if not m: raise HTTPException(status_code=404)
    db.delete(m); db.commit(); return {"message": "Deleted"}


# ─── Lien Waivers ─────────────────────────────────────────────────────────────

@router.get("/lien-waivers")
def list_lien_waivers(proj: Optional[Project] = Depends(_get_proj), db: Session = Depends(get_db)):
    if not proj: return []
    return [
        {"id": w.id, "draw_id": w.draw_id, "vendor_name": w.vendor_name or (w.subcontractor.name if w.subcontractor else None),
         "waiver_type": w.waiver_type, "amount": w.amount, "date_received": w.date_received,
         "notes": w.notes, "subcontractor_id": w.subcontractor_id}
        for w in db.query(LienWaiver).filter(LienWaiver.project_id == proj.id).order_by(LienWaiver.date_received.desc(), LienWaiver.id.desc()).all()
    ]


@router.post("/lien-waivers")
def create_lien_waiver(body: dict, proj: Project = Depends(_req_proj), db: Session = Depends(get_db)):
    if body.get("waiver_type") not in ("conditional", "unconditional"):
        raise HTTPException(status_code=400, detail="waiver_type must be conditional or unconditional")
    w = LienWaiver(project_id=proj.id, draw_id=body.get("draw_id"),
                   subcontractor_id=body.get("subcontractor_id"), vendor_name=body.get("vendor_name"),
                   waiver_type=body["waiver_type"], amount=body.get("amount"),
                   date_received=body.get("date_received"), notes=body.get("notes"))
    db.add(w); db.commit(); db.refresh(w)
    return {"id": w.id, "waiver_type": w.waiver_type}


@router.delete("/lien-waivers/{w_id}")
def delete_lien_waiver(w_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    w = (db.query(LienWaiver).join(Project, Project.id == LienWaiver.project_id)
         .filter(LienWaiver.id == w_id, Project.user_id == current_user.id).first())
    if not w: raise HTTPException(status_code=404)
    db.delete(w); db.commit(); return {"message": "Deleted"}


# ─── Project Documents ───────────────────────────────────────────────────────

_DOC_TYPES = {"contract","permit","rfi","submittal","drawing","report","other"}
_DOC_UPLOAD_DIR = os.path.join(os.path.dirname(__file__), "..", "..", "uploads", "docs")

DOC_ALLOWED_EXT = {".pdf",".jpg",".jpeg",".png",".docx",".doc",".xlsx",".xls",".pptx",".ppt",".txt"}


def _doc_out(doc):
    return {
        "id": doc.id, "doc_type": doc.doc_type, "title": doc.title,
        "original_filename": doc.original_filename, "external_url": doc.external_url,
        "notes": doc.notes, "draw_id": doc.draw_id, "category_id": doc.category_id,
        "has_file": bool(doc.file_path),
        "created_at": str(doc.created_at),
    }


@router.get("/documents")
def list_documents(doc_type: Optional[str] = None, proj: Optional[Project] = Depends(_get_proj), db: Session = Depends(get_db)):
    if not proj:
        return []
    q = db.query(ProjectDocument).filter(ProjectDocument.project_id == proj.id)
    if doc_type:
        q = q.filter(ProjectDocument.doc_type == doc_type)
    return [_doc_out(d) for d in q.order_by(ProjectDocument.created_at.desc()).all()]


@router.post("/documents/upload")
async def upload_document(
    file: UploadFile = File(None),
    title: str = Form(...),
    doc_type: str = Form("other"),
    notes: str = Form(None),
    draw_id: int = Form(None),
    category_id: int = Form(None),
    external_url: str = Form(None),
    proj: Project = Depends(_req_proj),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if doc_type not in _DOC_TYPES:
        raise HTTPException(status_code=400, detail=f"doc_type must be one of: {', '.join(_DOC_TYPES)}")
    import aiofiles
    file_path = None
    original_filename = None
    if file and file.filename:
        ext = Path(file.filename).suffix.lower()
        if ext not in DOC_ALLOWED_EXT:
            raise HTTPException(status_code=400, detail=f"Unsupported file type. Allowed: {', '.join(DOC_ALLOWED_EXT)}")
        os.makedirs(_DOC_UPLOAD_DIR, exist_ok=True)
        unique_name = f"{proj.id}_{uuid.uuid4().hex}{ext}"
        file_path = os.path.join(_DOC_UPLOAD_DIR, unique_name)
        content = await file.read()
        with open(file_path, "wb") as f_out:
            f_out.write(content)
        original_filename = file.filename

    doc = ProjectDocument(
        project_id=proj.id, user_id=current_user.id,
        doc_type=doc_type, title=title, file_path=file_path,
        original_filename=original_filename, external_url=external_url,
        notes=notes, draw_id=draw_id, category_id=category_id,
    )
    db.add(doc); db.commit(); db.refresh(doc)
    return _doc_out(doc)


@router.get("/documents/{doc_id}/file")
async def download_document(doc_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    doc = (db.query(ProjectDocument).join(Project, Project.id == ProjectDocument.project_id)
           .filter(ProjectDocument.id == doc_id, Project.user_id == current_user.id).first())
    if not doc or not doc.file_path:
        raise HTTPException(status_code=404)
    if not os.path.isfile(doc.file_path):
        raise HTTPException(status_code=404, detail="File not found on disk")
    from fastapi.responses import FileResponse as _FR
    return _FR(doc.file_path, filename=doc.original_filename or Path(doc.file_path).name)


@router.delete("/documents/{doc_id}")
def delete_document(doc_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    doc = (db.query(ProjectDocument).join(Project, Project.id == ProjectDocument.project_id)
           .filter(ProjectDocument.id == doc_id, Project.user_id == current_user.id).first())
    if not doc:
        raise HTTPException(status_code=404)
    if doc.file_path and os.path.isfile(doc.file_path):
        os.remove(doc.file_path)
    db.delete(doc); db.commit()
    return {"message": "Deleted"}


# ─── Portfolio Rollup ────────────────────────────────────────────────────────

@router.get("/portfolio")
def portfolio_rollup(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    """High-level summary across ALL projects for the current user."""
    projects = db.query(Project).filter(Project.user_id == current_user.id).order_by(Project.created_at).all()
    result = []
    for proj in projects:
        # Cost categories
        cats = db.query(CostCategory).filter(CostCategory.project_id == proj.id).all()
        total_budget = sum(c.budget for c in cats)
        total_invoiced = 0.0
        total_paid = 0.0
        for cat in cats:
            invoiced = db.query(func.coalesce(func.sum(InvoiceAllocation.amount), 0.0)).filter(InvoiceAllocation.category_id == cat.id).scalar() or 0
            total_invoiced += invoiced
            for a in db.query(InvoiceAllocation).filter(InvoiceAllocation.category_id == cat.id).all():
                inv = db.query(Invoice).filter(Invoice.id == a.invoice_id).first()
                if inv:
                    total_paid += (inv.amount_paid or 0) * (a.percentage / 100)
        # Draw status
        draws = db.query(Draw).filter(Draw.project_id == proj.id).all()
        # Pending approvals — scoped to this project
        pending_approvals = db.query(Invoice).filter(
            Invoice.user_id == current_user.id,
            Invoice.project_id == proj.id,
            Invoice.approval_status == "pending",
            Invoice.status == "processed",
        ).count()
        pct_burn = round((total_invoiced / total_budget * 100) if total_budget else 0, 1)
        result.append({
            "id": proj.id, "name": proj.name, "code": proj.code, "client": proj.client,
            "total_budget": total_budget,
            "total_invoiced": round(total_invoiced, 2),
            "total_paid": round(total_paid, 2),
            "total_remaining": round(total_budget - total_invoiced, 2),
            "pct_burn": pct_burn,
            "draw_count": len(draws),
            "pending_approvals": pending_approvals,
            "start_date": proj.start_date,
            "end_date": proj.end_date,
        })
    return {
        "projects": result,
        "totals": {
            "budget": round(sum(r["total_budget"] for r in result), 2),
            "invoiced": round(sum(r["total_invoiced"] for r in result), 2),
            "paid": round(sum(r["total_paid"] for r in result), 2),
            "remaining": round(sum(r["total_remaining"] for r in result), 2),
        },
    }


# ─── Subcontractor Directory ─────────────────────────────────────────────────

def _sub_out(s, today_str: str):
    def _expiry_flag(d):
        if not d: return "missing"
        return "expired" if d < today_str else ("expiring_soon" if d <= today_str[:4] + "-" + str(int(today_str[5:7]) + 2).zfill(2) + "-" + today_str[8:] else "ok")
    return {
        "id": s.id, "name": s.name, "trade": s.trade,
        "contact_name": s.contact_name, "contact_email": s.contact_email, "contact_phone": s.contact_phone,
        "contract_value": s.contract_value, "status": s.status,
        "insurance_expiry": s.insurance_expiry, "insurance_flag": _expiry_flag(s.insurance_expiry),
        "wsib_expiry": s.wsib_expiry, "wsib_flag": _expiry_flag(s.wsib_expiry),
        "notes": s.notes, "created_at": str(s.created_at),
    }


@router.get("/subcontractors")
def list_subcontractors(proj: Optional[Project] = Depends(_get_proj), db: Session = Depends(get_db)):
    if not proj:
        return []
    from datetime import datetime as _dt
    today = _dt.utcnow().strftime("%Y-%m-%d")
    return [_sub_out(s, today) for s in
            db.query(Subcontractor).filter(Subcontractor.project_id == proj.id)
            .order_by(Subcontractor.name).all()]


@router.post("/subcontractors")
def create_subcontractor(body: dict, proj: Project = Depends(_req_proj), db: Session = Depends(get_db)):
    if not body.get("name"):
        raise HTTPException(status_code=400, detail="name is required")
    s = Subcontractor(
        project_id=proj.id,
        name=body["name"], trade=body.get("trade"),
        contact_name=body.get("contact_name"), contact_email=body.get("contact_email"), contact_phone=body.get("contact_phone"),
        contract_value=body.get("contract_value"), status=body.get("status", "active"),
        insurance_expiry=body.get("insurance_expiry"), wsib_expiry=body.get("wsib_expiry"),
        notes=body.get("notes"),
    )
    db.add(s); db.commit(); db.refresh(s)
    from datetime import datetime as _dt
    return _sub_out(s, _dt.utcnow().strftime("%Y-%m-%d"))


@router.put("/subcontractors/{sub_id}")
def update_subcontractor(sub_id: int, body: dict, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    s = (db.query(Subcontractor).join(Project, Project.id == Subcontractor.project_id)
         .filter(Subcontractor.id == sub_id, Project.user_id == current_user.id).first())
    if not s: raise HTTPException(status_code=404)
    for f in ("name","trade","contact_name","contact_email","contact_phone","contract_value","status","insurance_expiry","wsib_expiry","notes"):
        if f in body: setattr(s, f, body[f])
    db.commit(); db.refresh(s)
    from datetime import datetime as _dt
    return _sub_out(s, _dt.utcnow().strftime("%Y-%m-%d"))


@router.delete("/subcontractors/{sub_id}")
def delete_subcontractor(sub_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    s = (db.query(Subcontractor).join(Project, Project.id == Subcontractor.project_id)
         .filter(Subcontractor.id == sub_id, Project.user_id == current_user.id).first())
    if not s: raise HTTPException(status_code=404)
    db.delete(s); db.commit()
    return {"message": "Deleted"}


# ─── Committed Costs (POs / Contracts) ───────────────────────────────────────

def _cc_out(cc, db):
    return {
        "id": cc.id, "vendor": cc.vendor, "description": cc.description,
        "contract_amount": cc.contract_amount, "invoiced_to_date": cc.invoiced_to_date or 0,
        "remaining_to_invoice": round(cc.contract_amount - (cc.invoiced_to_date or 0), 2),
        "status": cc.status, "contract_date": cc.contract_date,
        "expected_completion": cc.expected_completion, "notes": cc.notes,
        "category_id": cc.category_id,
        "category_name": cc.category.name if cc.category else None,
        "created_at": str(cc.created_at),
    }


@router.get("/committed-costs")
def list_committed_costs(proj: Optional[Project] = Depends(_get_proj), db: Session = Depends(get_db)):
    if not proj:
        return []
    return [_cc_out(cc, db) for cc in
            db.query(CommittedCost).filter(CommittedCost.project_id == proj.id)
            .order_by(CommittedCost.contract_date.desc(), CommittedCost.id.desc()).all()]


@router.post("/committed-costs")
def create_committed_cost(body: dict, proj: Project = Depends(_req_proj), db: Session = Depends(get_db)):
    if not body.get("vendor"):
        raise HTTPException(status_code=400, detail="vendor is required")
    if body.get("contract_amount") is None:
        raise HTTPException(status_code=400, detail="contract_amount is required")
    cat_id = body.get("category_id")
    if cat_id:
        cat = db.query(CostCategory).filter(CostCategory.id == cat_id, CostCategory.project_id == proj.id).first()
        if not cat:
            raise HTTPException(status_code=404, detail="Cost category not found in this project")
    cc = CommittedCost(
        project_id=proj.id, category_id=cat_id,
        vendor=body["vendor"], description=body.get("description"),
        contract_amount=float(body["contract_amount"]),
        invoiced_to_date=float(body.get("invoiced_to_date", 0)),
        status=body.get("status", "active"),
        contract_date=body.get("contract_date"),
        expected_completion=body.get("expected_completion"),
        notes=body.get("notes"),
    )
    db.add(cc)
    db.commit()
    db.refresh(cc)
    return _cc_out(cc, db)


@router.put("/committed-costs/{cc_id}")
def update_committed_cost(cc_id: int, body: dict, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    cc = (db.query(CommittedCost).join(Project, Project.id == CommittedCost.project_id)
          .filter(CommittedCost.id == cc_id, Project.user_id == current_user.id).first())
    if not cc:
        raise HTTPException(status_code=404)
    for field in ("vendor", "description", "contract_amount", "invoiced_to_date", "status", "contract_date", "expected_completion", "notes", "category_id"):
        if field in body:
            setattr(cc, field, body[field])
    db.commit()
    db.refresh(cc)
    return _cc_out(cc, db)


@router.delete("/committed-costs/{cc_id}")
def delete_committed_cost(cc_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    cc = (db.query(CommittedCost).join(Project, Project.id == CommittedCost.project_id)
          .filter(CommittedCost.id == cc_id, Project.user_id == current_user.id).first())
    if not cc:
        raise HTTPException(status_code=404)
    db.delete(cc)
    db.commit()
    return {"message": "Deleted"}


# ─── Lender Tokens ───────────────────────────────────────────────────────────

@router.get("/lender-tokens")
def list_lender_tokens(proj: Optional[Project] = Depends(_get_proj), db: Session = Depends(get_db)):
    if not proj:
        return []
    tokens = db.query(LenderToken).filter(LenderToken.project_id == proj.id).order_by(LenderToken.created_at.desc()).all()
    return [
        {"id": t.id, "label": t.label, "token": t.token, "draw_id": t.draw_id,
         "is_active": t.is_active, "expires_at": t.expires_at, "created_at": str(t.created_at)}
        for t in tokens
    ]


@router.post("/lender-tokens")
def create_lender_token(body: dict, proj: Project = Depends(_req_proj), db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    import secrets as _sec
    if not body.get("label"):
        raise HTTPException(status_code=400, detail="label is required")
    draw_id = body.get("draw_id")
    if draw_id:
        draw = db.query(Draw).filter(Draw.id == draw_id, Draw.project_id == proj.id).first()
        if not draw:
            raise HTTPException(status_code=404, detail="Draw not found in this project")
    token_str = _sec.token_urlsafe(24)  # 32 chars URL-safe
    t = LenderToken(
        project_id=proj.id, draw_id=draw_id,
        token=token_str, label=body["label"],
        created_by=current_user.id,
        is_active=True, expires_at=body.get("expires_at"),
    )
    db.add(t); db.commit(); db.refresh(t)
    return {"id": t.id, "label": t.label, "token": t.token, "draw_id": t.draw_id}


@router.put("/lender-tokens/{token_id}/toggle")
def toggle_lender_token(token_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    t = (db.query(LenderToken).join(Project, Project.id == LenderToken.project_id)
         .filter(LenderToken.id == token_id, Project.user_id == current_user.id).first())
    if not t:
        raise HTTPException(status_code=404)
    t.is_active = not t.is_active
    db.commit()
    return {"id": t.id, "is_active": t.is_active}


@router.delete("/lender-tokens/{token_id}")
def delete_lender_token(token_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    t = (db.query(LenderToken).join(Project, Project.id == LenderToken.project_id)
         .filter(LenderToken.id == token_id, Project.user_id == current_user.id).first())
    if not t:
        raise HTTPException(status_code=404)
    db.delete(t); db.commit()
    return {"message": "Deleted"}


# ─── Public Lender Package (no auth — token-gated) ───────────────────────────

_lender_router = APIRouter(prefix="/api/lender", tags=["lender"])


@_lender_router.get("/{token}")
def lender_package(token: str, db: Session = Depends(get_db)):
    """Public endpoint — returns a curated draw package for external lenders.
    No authentication required; access is controlled by the token secret.
    Sensitive internal data (markup %, govt claims, payroll) is excluded."""
    from datetime import datetime as _dt
    t = db.query(LenderToken).filter(LenderToken.token == token, LenderToken.is_active == True).first()
    if not t:
        raise HTTPException(status_code=404, detail="Link not found or has been deactivated.")
    # Check expiry
    if t.expires_at and t.expires_at < _dt.utcnow().strftime("%Y-%m-%d"):
        raise HTTPException(status_code=403, detail="This link has expired.")

    proj = db.query(Project).filter(Project.id == t.project_id).first()
    if not proj:
        raise HTTPException(status_code=404)

    # Determine which draws to include
    if t.draw_id:
        draws = db.query(Draw).filter(Draw.id == t.draw_id, Draw.project_id == proj.id).all()
    else:
        draws = db.query(Draw).filter(Draw.project_id == proj.id).order_by(Draw.draw_number).all()

    draws_out = []
    all_invoice_ids = set()
    for draw in draws:
        invs = db.query(Invoice).filter(Invoice.draw_id == draw.id, Invoice.status == "processed").all()
        all_invoice_ids.update(i.id for i in invs)
        draw_invoiced = sum(i.total_due or 0 for i in invs)
        draw_lender_sub = sum(i.lender_submitted_amt or 0 for i in invs)
        draw_lender_app = sum(i.lender_approved_amt or 0 for i in invs)
        inv_rows = []
        for i in invs:
            holdback = round((i.subtotal or i.total_due or 0) * (i.holdback_pct or 0) / 100, 2)
            inv_rows.append({
                "invoice_number": i.invoice_number, "vendor": i.vendor_name,
                "date": i.invoice_date, "subtotal": i.subtotal, "tax": i.tax_total,
                "total": i.total_due, "holdback_pct": i.holdback_pct, "holdback_amt": holdback,
                "lender_submitted": i.lender_submitted_amt, "lender_approved": i.lender_approved_amt,
                "lender_status": i.lender_status, "payment_status": i.payment_status,
                "approval_status": i.approval_status,
            })
        draws_out.append({
            "draw_number": draw.draw_number, "status": draw.status,
            "submission_date": draw.submission_date, "notes": draw.notes,
            "total_invoiced": round(draw_invoiced, 2),
            "total_lender_submitted": round(draw_lender_sub, 2),
            "total_lender_approved": round(draw_lender_app, 2),
            "invoice_count": len(invs),
            "invoices": inv_rows,
        })

    # Category breakdown (curated — no internal rates)
    categories = db.query(CostCategory).filter(CostCategory.project_id == proj.id).order_by(CostCategory.display_order).all()
    cat_rows = []
    for cat in categories:
        invoiced = db.query(func.coalesce(func.sum(InvoiceAllocation.amount), 0.0)).filter(
            InvoiceAllocation.category_id == cat.id
        ).scalar() or 0
        lender_app = 0.0
        for a in db.query(InvoiceAllocation).filter(InvoiceAllocation.category_id == cat.id).all():
            inv = db.query(Invoice).filter(Invoice.id == a.invoice_id).first()
            if inv:
                lender_app += (inv.lender_approved_amt or 0) * (a.percentage / 100)
        if invoiced > 0:
            cat_rows.append({
                "name": cat.name,
                "budget": cat.lender_budget or cat.budget,   # lender sees lender budget if set
                "lender_budget": cat.lender_budget,
                "invoiced": round(invoiced, 2), "lender_approved": round(lender_app, 2),
            })

    # Totals
    all_invs = db.query(Invoice).filter(Invoice.draw_id.in_(d.id for d in draws), Invoice.status == "processed").all() if draws else []
    holdback_held = sum(round((i.subtotal or i.total_due or 0) * (i.holdback_pct or 0) / 100, 2)
                        for i in all_invs if not i.holdback_released)

    return {
        "project": {"name": proj.name, "code": proj.code, "client": proj.client, "address": proj.address},
        "label": t.label, "generated_at": _dt.utcnow().strftime("%Y-%m-%d"),
        "draws": draws_out,
        "categories": cat_rows,
        "summary": {
            "total_invoiced": round(sum(d["total_invoiced"] for d in draws_out), 2),
            "total_lender_submitted": round(sum(d["total_lender_submitted"] for d in draws_out), 2),
            "total_lender_approved": round(sum(d["total_lender_approved"] for d in draws_out), 2),
            "holdback_held": round(holdback_held, 2),
            "draw_count": len(draws_out),
            "invoice_count": sum(d["invoice_count"] for d in draws_out),
        },
    }


# ─── Aged Payables ───────────────────────────────────────────────────────────

@router.get("/aged-payables")
def aged_payables(
    proj: Optional[Project] = Depends(_get_proj),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Per-vendor aged payables: unpaid/partially-paid invoices grouped by ageing bucket."""
    from datetime import datetime as _dt
    from collections import defaultdict
    today = _dt.utcnow().strftime("%Y-%m-%d")

    # Build project-scoped filter for aged payables
    _ap_draws = db.query(Draw).filter(Draw.project_id == proj.id).all() if proj else []
    _ap_claims = db.query(Claim).filter(Claim.project_id == proj.id).all() if proj else []
    _ap_draw_ids = [d.id for d in _ap_draws]
    _ap_claim_ids = [c.id for c in _ap_claims]
    from sqlalchemy import or_ as _ap_or
    _ap_conds = [Invoice.project_id == proj.id] if proj else [Invoice.user_id == current_user.id]
    if proj and _ap_draw_ids:
        _ap_conds.append(Invoice.draw_id.in_(_ap_draw_ids))
    if proj and _ap_claim_ids:
        _ap_conds.append(Invoice.provincial_claim_id.in_(_ap_claim_ids))
        _ap_conds.append(Invoice.federal_claim_id.in_(_ap_claim_ids))

    unpaid = db.query(Invoice).filter(
        Invoice.user_id == current_user.id,
        Invoice.status == "processed",
        Invoice.payment_status != "paid",
        _ap_or(*_ap_conds),
    ).all()

    # Group by vendor, bucket by days past due (using due_date or invoice_date)
    vendor_map: dict = defaultdict(lambda: {"current": 0, "over_30": 0, "over_60": 0, "over_90": 0, "invoices": []})

    for inv in unpaid:
        vendor = inv.vendor_name or "Unknown"
        due = inv.due_date or inv.invoice_date
        outstanding = (inv.total_due or 0) - (inv.amount_paid or 0)
        if outstanding <= 0:
            continue
        days = 0
        if due:
            try:
                days = (_dt.strptime(today, "%Y-%m-%d") - _dt.strptime(due, "%Y-%m-%d")).days
            except ValueError:
                days = 0
        bucket = "over_90" if days > 90 else "over_60" if days > 60 else "over_30" if days > 30 else "current"
        vendor_map[vendor][bucket] += outstanding
        vendor_map[vendor]["invoices"].append({
            "id": inv.id, "invoice_number": inv.invoice_number, "date": inv.invoice_date,
            "due_date": inv.due_date, "total": inv.total_due, "paid": inv.amount_paid,
            "outstanding": round(outstanding, 2), "days_past_due": max(days, 0),
            "bucket": bucket, "payment_status": inv.payment_status,
        })

    rows = []
    for vendor, data in sorted(vendor_map.items()):
        total = sum(data[b] for b in ("current","over_30","over_60","over_90"))
        rows.append({
            "vendor": vendor,
            "current": round(data["current"], 2),
            "over_30": round(data["over_30"], 2),
            "over_60": round(data["over_60"], 2),
            "over_90": round(data["over_90"], 2),
            "total": round(total, 2),
            "invoice_count": len(data["invoices"]),
            "invoices": sorted(data["invoices"], key=lambda i: i["days_past_due"], reverse=True),
        })
    rows.sort(key=lambda r: r["total"], reverse=True)

    totals = {
        "current": round(sum(r["current"] for r in rows), 2),
        "over_30": round(sum(r["over_30"] for r in rows), 2),
        "over_60": round(sum(r["over_60"] for r in rows), 2),
        "over_90": round(sum(r["over_90"] for r in rows), 2),
        "total": round(sum(r["total"] for r in rows), 2),
    }
    return {"vendors": rows, "totals": totals, "as_of": today}


# ─── Accounting Export ───────────────────────────────────────────────────────

@router.get("/export/accounting-csv")
def export_accounting_csv(
    format: Optional[str] = "qbo",   # qbo | xero
    proj: Optional[Project] = Depends(_get_proj),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Export invoices as CSV shaped for QuickBooks Online (qbo) or Xero import.
    Columns follow the standard Accounts Payable / Bills import format."""
    import csv, io as _io
    from datetime import datetime as _dt

    # Project-scoped: only export invoices for this project
    _export_proj_id = proj.id if proj else None
    _inv_q = db.query(Invoice).filter(
        Invoice.user_id == current_user.id,
        Invoice.status == "processed",
    )
    if _export_proj_id:
        _inv_q = _inv_q.filter(Invoice.project_id == _export_proj_id)
    invoices = _inv_q.order_by(Invoice.invoice_date.desc()).all()

    buf = _io.StringIO()
    today = _dt.utcnow().strftime("%Y-%m-%d")

    if format == "xero":
        # Xero Bills import format
        w = csv.writer(buf)
        w.writerow(["*ContactName","*InvoiceNumber","*InvoiceDate","*DueDate","*Description",
                    "*Quantity","*UnitAmount","*AccountCode","*TaxType","Currency","TrackingName1"])
        for inv in invoices:
            w.writerow([
                inv.vendor_name or "Unknown",
                inv.invoice_number or "",
                inv.invoice_date or today,
                inv.due_date or inv.invoice_date or today,
                f"Invoice from {inv.vendor_name or 'vendor'}",
                1,
                inv.subtotal or inv.total_due or 0,
                "200",   # standard AP account
                "INPUT2",  # HST on purchases
                inv.currency or "CAD",
                proj.name if proj else "",
            ])
    else:
        # QuickBooks Online Bills import format
        w = csv.writer(buf)
        w.writerow(["Vendor","Bill No","Date","Due Date","Memo","Category",
                    "Description","Amount","Tax Amount","Total","Currency","Reference"])
        for inv in invoices:
            # Find primary category from allocations
            cat_name = ""
            alloc = db.query(InvoiceAllocation).filter(InvoiceAllocation.invoice_id == inv.id).first()
            if alloc and alloc.category:
                cat_name = alloc.category.name
            w.writerow([
                inv.vendor_name or "",
                inv.invoice_number or "",
                inv.invoice_date or today,
                inv.due_date or inv.invoice_date or today,
                f"Project: {proj.name}" if proj else "",
                cat_name,
                f"Invoice {inv.invoice_number or inv.id}",
                inv.subtotal or (inv.total_due or 0) - (inv.tax_total or 0),
                inv.tax_total or 0,
                inv.total_due or 0,
                inv.currency or "CAD",
                inv.vendor_on_record or "",
            ])

    csv_bytes = buf.getvalue().encode("utf-8-sig")  # BOM for Excel compatibility
    filename = f"invoices_{format}_{today}.csv"
    return StreamingResponse(
        iter([csv_bytes]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ─── Cash Flow Projection ────────────────────────────────────────────────────

@router.get("/cash-flow")
def cash_flow(
    proj: Optional[Project] = Depends(_get_proj),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Month-by-month cash flow: actual spend (invoiced), actual paid, draw receipts, and projected future spend from committed costs."""
    if not proj:
        return {"months": []}

    from collections import defaultdict
    from datetime import datetime as dt

    spend: dict = defaultdict(float)     # invoiced by invoice_date month
    paid: dict = defaultdict(float)      # actual payments by payment_date month
    receipts: dict = defaultdict(float)  # draw lender_approved by submission_date month
    projected: dict = defaultdict(float) # committed costs by expected_completion month

    # Actual spend — project-scoped invoices by invoice_date
    proj_draws_cf = db.query(Draw).filter(Draw.project_id == proj.id).all()
    proj_draw_ids_cf = [d.id for d in proj_draws_cf]
    proj_claims_cf = db.query(Claim).filter(Claim.project_id == proj.id).all()
    proj_claim_ids_cf = [c.id for c in proj_claims_cf]
    _cf_conds = [Invoice.project_id == proj.id]
    if proj_draw_ids_cf:
        _cf_conds.append(Invoice.draw_id.in_(proj_draw_ids_cf))
    if proj_claim_ids_cf:
        _cf_conds.append(Invoice.provincial_claim_id.in_(proj_claim_ids_cf))
        _cf_conds.append(Invoice.federal_claim_id.in_(proj_claim_ids_cf))
    from sqlalchemy import or_ as _or2_
    _cf_inv = db.query(Invoice).filter(
        Invoice.user_id == current_user.id,
        Invoice.status == "processed",
        _or2_(*_cf_conds),
    ).all()
    for inv in _cf_inv:
        date_str = inv.invoice_date or (str(inv.processed_at)[:10] if inv.processed_at else None)
        if date_str and len(date_str) >= 7:
            m = date_str[:7]
            spend[m] += inv.total_due or 0

    # Actual payments by payment date (project-scoped)
    _cf_inv_ids = {i.id for i in _cf_inv}
    for payment in db.query(Payment).join(Invoice, Invoice.id == Payment.invoice_id).filter(
        Invoice.user_id == current_user.id,
        Invoice.id.in_(_cf_inv_ids) if _cf_inv_ids else Invoice.id == -1,
    ).all():
        if payment.payment_date and len(payment.payment_date) >= 7:
            m = payment.payment_date[:7]
            paid[m] += payment.amount or 0

    # Draw receipts — lender-approved amounts by draw submission_date
    for draw in db.query(Draw).filter(Draw.project_id == proj.id).all():
        if draw.submission_date and len(draw.submission_date) >= 7:
            m = draw.submission_date[:7]
            # Sum lender_approved_amt for all invoices in this draw
            draw_total = db.query(func.coalesce(func.sum(Invoice.lender_approved_amt), 0.0)).filter(
                Invoice.draw_id == draw.id, Invoice.user_id == current_user.id
            ).scalar() or 0
            receipts[m] += draw_total

    # Projected future spend from active committed costs by expected_completion
    today_m = dt.utcnow().strftime("%Y-%m")
    for cc in db.query(CommittedCost).filter(CommittedCost.project_id == proj.id, CommittedCost.status == "active").all():
        if cc.expected_completion and len(cc.expected_completion) >= 7:
            m = cc.expected_completion[:7]
            if m >= today_m:  # only future months
                remaining = cc.contract_amount - (cc.invoiced_to_date or 0)
                if remaining > 0:
                    projected[m] += remaining

    # Merge all months, sort chronologically
    all_months = sorted(set(list(spend.keys()) + list(paid.keys()) + list(receipts.keys()) + list(projected.keys())))
    if not all_months:
        return {"months": []}

    cumulative = 0.0
    result = []
    for m in all_months:
        s = round(spend.get(m, 0), 2)
        p = round(paid.get(m, 0), 2)
        r = round(receipts.get(m, 0), 2)
        pr = round(projected.get(m, 0), 2)
        net = round(r - s, 2)
        cumulative = round(cumulative + net, 2)
        result.append({
            "month": m,
            "invoiced": s,
            "paid": p,
            "draw_receipts": r,
            "projected_spend": pr,
            "net": net,
            "cumulative": cumulative,
        })

    return {
        "months": result,
        "totals": {
            "invoiced": round(sum(m["invoiced"] for m in result), 2),
            "paid": round(sum(m["paid"] for m in result), 2),
            "draw_receipts": round(sum(m["draw_receipts"] for m in result), 2),
            "projected_spend": round(sum(m["projected_spend"] for m in result), 2),
        },
    }


# ─── Dashboard ────────────────────────────────────────────────────────────────

@router.get("/dashboard")
def project_dashboard(proj: Optional[Project] = Depends(_get_proj), db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    if not proj:
        return {"project": None}

    categories = (
        db.query(CostCategory)
        .filter(CostCategory.project_id == proj.id)
        .order_by(CostCategory.display_order)
        .all()
    )

    subdivisions = (
        db.query(SubDivision)
        .filter(SubDivision.project_id == proj.id)
        .order_by(SubDivision.display_order)
        .all()
    )

    # Pre-load active committed costs keyed by category_id
    active_ccs = db.query(CommittedCost).filter(
        CommittedCost.project_id == proj.id, CommittedCost.status == "active"
    ).all()
    cc_by_cat: dict = {}
    for cc in active_ccs:
        if cc.category_id:
            cc_by_cat[cc.category_id] = cc_by_cat.get(cc.category_id, 0.0) + cc.contract_amount

    # Pre-load all approved change orders for this project keyed by category_id
    approved_cos = db.query(ChangeOrder).filter(
        ChangeOrder.project_id == proj.id, ChangeOrder.status == "approved"
    ).all()
    co_by_cat: dict = {}
    co_project_level = 0.0
    for co in approved_cos:
        if co.category_id:
            co_by_cat[co.category_id] = co_by_cat.get(co.category_id, 0.0) + co.amount
        else:
            co_project_level += co.amount

    # Build category summary
    cat_summary = []
    for cat in categories:
        # Total allocated to this category
        alloc_sum = (
            db.query(func.coalesce(func.sum(InvoiceAllocation.amount), 0.0))
            .filter(InvoiceAllocation.category_id == cat.id)
            .scalar()
        )
        # Total paid for invoices allocated to this category
        paid_sum = 0.0
        alloc_rows = db.query(InvoiceAllocation).filter(InvoiceAllocation.category_id == cat.id).all()
        for a in alloc_rows:
            inv = db.query(Invoice).filter(Invoice.id == a.invoice_id).first()
            if inv:
                paid_sum += (inv.amount_paid or 0.0) * (a.percentage / 100.0)

        # Sub-category invoiced/paid amounts
        sc_data = []
        for sc in cat.sub_categories:
            sc_invoiced = (
                db.query(func.coalesce(func.sum(InvoiceAllocation.amount), 0.0))
                .filter(InvoiceAllocation.category_id == cat.id, InvoiceAllocation.sub_category_id == sc.id)
                .scalar()
            )
            sc_paid = sum(
                (db.query(Invoice).filter(Invoice.id == a.invoice_id).first().amount_paid or 0.0) * (a.percentage / 100.0)
                for a in db.query(InvoiceAllocation).filter(
                    InvoiceAllocation.category_id == cat.id, InvoiceAllocation.sub_category_id == sc.id
                ).all()
                if db.query(Invoice).filter(Invoice.id == a.invoice_id).first()
            )
            sc_data.append({
                "id": sc.id, "name": sc.name, "budget": sc.budget or 0,
                "invoiced": round(sc_invoiced, 2), "paid": round(sc_paid, 2),
                "remaining": round((sc.budget or 0) - sc_invoiced, 2),
            })

        co_adj = co_by_cat.get(cat.id, 0.0)
        committed = round(cc_by_cat.get(cat.id, 0.0), 2)
        revised_budget = cat.budget + co_adj
        # Lender budget: explicit lender_budget field, or fall back to internal revised budget
        lender_budget = (cat.lender_budget or cat.budget) + co_adj
        pct_burn = round((alloc_sum / revised_budget * 100) if revised_budget else 0, 1)
        lender_pct_burn = round((alloc_sum / lender_budget * 100) if lender_budget else 0, 1)
        cat_data = {
            "id": cat.id,
            "name": cat.name,
            "budget": cat.budget,
            "lender_budget_raw": cat.lender_budget,       # None means "same as internal"
            "co_adjustment": round(co_adj, 2),
            "revised_budget": round(revised_budget, 2),
            "lender_budget": round(lender_budget, 2),     # lender view: lender_budget_raw + CO
            "committed": committed,
            "exposed": round(committed - alloc_sum, 2),
            "invoiced": round(alloc_sum, 2),
            "paid": round(paid_sum, 2),
            "remaining": round(revised_budget - alloc_sum, 2),
            "lender_remaining": round(lender_budget - alloc_sum, 2),
            "pct_burn": pct_burn,
            "lender_pct_burn": lender_pct_burn,
            "is_per_subdivision": cat.is_per_subdivision,
            "sub_categories": sc_data,
        }

        # Per-subdivision breakdown for Fiber Build
        if cat.is_per_subdivision:
            sd_data = []
            for sd in subdivisions:
                sd_budget_row = db.query(SubDivisionBudget).filter(
                    SubDivisionBudget.category_id == cat.id,
                    SubDivisionBudget.subdivision_id == sd.id,
                ).first()
                sd_budget = sd_budget_row.budget if sd_budget_row else 0.0
                sd_invoiced = (
                    db.query(func.coalesce(func.sum(InvoiceAllocation.amount), 0.0))
                    .filter(
                        InvoiceAllocation.category_id == cat.id,
                        InvoiceAllocation.subdivision_id == sd.id,
                    )
                    .scalar()
                )
                sd_data.append({
                    "subdivision_id": sd.id,
                    "name": sd.name,
                    "budget": sd_budget,
                    "invoiced": round(sd_invoiced, 2),
                    "remaining": round(sd_budget - sd_invoiced, 2),
                })
            cat_data["subdivisions"] = sd_data

        cat_summary.append(cat_data)

    # Overall totals (use revised budget = original + approved COs)
    total_budget = sum(c["budget"] for c in cat_summary)
    total_co_adjustment = sum(c["co_adjustment"] for c in cat_summary) + co_project_level
    total_revised_budget = total_budget + total_co_adjustment
    # Lender total budget: use project.lender_budget if set, else sum category lender budgets
    total_lender_budget_cats = sum(c["lender_budget"] for c in cat_summary)
    total_lender_budget = (proj.lender_budget + total_co_adjustment) if proj.lender_budget else total_lender_budget_cats
    total_committed = round(sum(cc.contract_amount for cc in active_ccs), 2)
    total_invoiced = sum(c["invoiced"] for c in cat_summary)
    total_paid = sum(c["paid"] for c in cat_summary)

    # Draws + claims (needed before scoped invoice filter)
    draws = db.query(Draw).filter(Draw.project_id == proj.id).order_by(Draw.draw_number).all()
    draws_summary = [_draw_out(d, db).model_dump() for d in draws]
    prov_claims = db.query(Claim).filter(Claim.project_id == proj.id, Claim.claim_type == "provincial").order_by(Claim.claim_number).all()
    fed_claims  = db.query(Claim).filter(Claim.project_id == proj.id, Claim.claim_type == "federal").order_by(Claim.claim_number).all()

    # Project-scoped invoice filter: invoices tagged to this project OR assigned to its draws/claims.
    # This prevents multi-project contamination in all aggregate queries below.
    _draw_ids  = [d.id for d in draws]
    _claim_ids = [c.id for c in prov_claims + fed_claims]
    _proj_inv_conditions = [Invoice.project_id == proj.id]
    if _draw_ids:
        _proj_inv_conditions.append(Invoice.draw_id.in_(_draw_ids))
    if _claim_ids:
        _proj_inv_conditions.append(Invoice.provincial_claim_id.in_(_claim_ids))
        _proj_inv_conditions.append(Invoice.federal_claim_id.in_(_claim_ids))
    from sqlalchemy import or_ as _or_
    def _proj_inv_base():
        return db.query(Invoice).filter(
            Invoice.user_id == current_user.id,
            Invoice.status == "processed",
            _or_(*_proj_inv_conditions),
        )

    # Invoice counts (project-scoped)
    all_invoices = _proj_inv_base().all()
    unallocated = 0
    for inv in all_invoices:
        has_alloc = db.query(InvoiceAllocation).filter(InvoiceAllocation.invoice_id == inv.id).count()
        if not has_alloc:
            unallocated += 1

    # Aging buckets (project-scoped)
    from datetime import datetime, timedelta
    today = datetime.utcnow().strftime("%Y-%m-%d")
    unpaid = _proj_inv_base().filter(Invoice.payment_status != "paid").all()
    aging = {"current": 0, "over_30": 0, "over_60": 0, "over_90": 0}
    for inv in unpaid:
        due = inv.due_date or inv.invoice_date
        if not due:
            aging["current"] += (inv.total_due or 0.0)
            continue
        try:
            days = (datetime.strptime(today, "%Y-%m-%d") - datetime.strptime(due, "%Y-%m-%d")).days
        except ValueError:
            days = 0
        amt = (inv.total_due or 0.0) - (inv.amount_paid or 0.0)
        if days > 90:
            aging["over_90"] += amt
        elif days > 60:
            aging["over_60"] += amt
        elif days > 30:
            aging["over_30"] += amt
        else:
            aging["current"] += amt

    # Unassigned to draws/claims (project-scoped)
    no_draw = _proj_inv_base().filter(Invoice.draw_id.is_(None)).count()
    no_prov = _proj_inv_base().filter(Invoice.provincial_claim_id.is_(None)).count()
    no_fed  = _proj_inv_base().filter(Invoice.federal_claim_id.is_(None)).count()

    # Cost tracking summary (project-scoped)
    all_processed = _proj_inv_base().all()
    committed_total = sum(i.received_total or i.total_due or 0 for i in all_processed)
    lender_approved = sum(i.lender_approved_amt or 0 for i in all_processed)
    lender_pending  = sum(i.lender_submitted_amt or 0 for i in all_processed if i.lender_status == "pending")
    lender_rejected = sum(i.lender_submitted_amt or 0 for i in all_processed if i.lender_status == "rejected")
    govt_approved   = sum(i.govt_approved_amt or 0 for i in all_processed)
    govt_pending    = sum(i.govt_submitted_amt or 0 for i in all_processed if i.govt_status == "pending")
    govt_rejected   = sum(i.govt_submitted_amt or 0 for i in all_processed if i.govt_status == "rejected")

    # Payroll summary (project-scoped via project_id)
    payroll_entries = db.query(PayrollEntry).filter(
        PayrollEntry.user_id == current_user.id,
        PayrollEntry.project_id == proj.id,
        PayrollEntry.status == "processed",
    ).all()
    payroll_committed      = sum(p.gross_pay or 0 for p in payroll_entries)
    payroll_lender_approved = sum(p.lender_approved_amt or 0 for p in payroll_entries)
    payroll_govt_approved   = sum(p.govt_approved_amt or 0 for p in payroll_entries)

    # Holdback aggregates (project-scoped)
    holdback_invoices = _proj_inv_base().filter(Invoice.holdback_pct > 0).all()
    holdback_held = sum(round((inv.subtotal or inv.total_due or 0) * (inv.holdback_pct or 0) / 100, 2)
                        for inv in holdback_invoices if not inv.holdback_released)
    holdback_released_total = sum(round((inv.subtotal or inv.total_due or 0) * (inv.holdback_pct or 0) / 100, 2)
                                   for inv in holdback_invoices if inv.holdback_released)
    holdback_total = holdback_held + holdback_released_total

    # Approval summary (project-scoped)
    approval_counts = {
        "pending":  _proj_inv_base().filter(Invoice.approval_status == "pending").count(),
        "approved": _proj_inv_base().filter(Invoice.approval_status == "approved").count(),
        "rejected": _proj_inv_base().filter(Invoice.approval_status == "rejected").count(),
    }

    # All change orders (all statuses) for the CO log
    all_cos = db.query(ChangeOrder).filter(ChangeOrder.project_id == proj.id)\
        .order_by(ChangeOrder.date.desc(), ChangeOrder.id.desc()).all()

    # All committed costs for the panel
    all_ccs = db.query(CommittedCost).filter(CommittedCost.project_id == proj.id)\
        .order_by(CommittedCost.contract_date.desc(), CommittedCost.id.desc()).all()

    return {
        "project": ProjectOut.model_validate(proj).model_dump(),
        "total_budget": total_budget,
        "total_lender_budget": round(total_lender_budget, 2),
        "total_co_adjustment": round(total_co_adjustment, 2),
        "total_revised_budget": round(total_revised_budget, 2),
        "total_committed": total_committed,
        "total_invoiced": round(total_invoiced, 2),
        "total_paid": round(total_paid, 2),
        "total_remaining": round(total_revised_budget - total_invoiced, 2),
        "total_lender_remaining": round(total_lender_budget - total_invoiced, 2),
        "categories": cat_summary,
        "change_orders": [
            {"id": co.id, "co_number": co.co_number, "description": co.description,
             "amount": co.amount, "status": co.status, "date": co.date,
             "issued_by": co.issued_by, "category_id": co.category_id,
             "category_name": co.category.name if co.category else None}
            for co in all_cos
        ],
        "committed_costs": [_cc_out(cc, db) for cc in all_ccs],
        "holdback": {
            "held": round(holdback_held, 2),
            "released": round(holdback_released_total, 2),
            "total": round(holdback_total, 2),
            "invoice_count": len(holdback_invoices),
            "unreleased_count": sum(1 for i in holdback_invoices if not i.holdback_released),
        },
        "approval": approval_counts,
        "unallocated_invoices": unallocated,
        "aging": {k: round(v, 2) for k, v in aging.items()},
        "draws": draws_summary,
        "provincial_claims": [_claim_out(c, db).model_dump() for c in prov_claims],
        "federal_claims": [_claim_out(c, db).model_dump() for c in fed_claims],
        "invoices_without_draw": no_draw,
        "invoices_without_claim": no_prov + no_fed,
        # Cost tracking 4-view
        "cost_tracking": {
            "committed": round(committed_total + payroll_committed, 2),
            "lender": {
                "approved": round(lender_approved + payroll_lender_approved, 2),
                "pending": round(lender_pending, 2),
                "rejected": round(lender_rejected, 2),
            },
            "govt": {
                "approved": round(govt_approved + payroll_govt_approved, 2),
                "pending": round(govt_pending, 2),
                "rejected": round(govt_rejected, 2),
            },
            "net_position": {
                "committed": round(committed_total + payroll_committed, 2),
                "recovered_lender": round(lender_approved + payroll_lender_approved, 2),
                "recovered_govt": round(govt_approved + payroll_govt_approved, 2),
                "out_of_pocket": round((committed_total + payroll_committed) - (lender_approved + payroll_lender_approved) - (govt_approved + payroll_govt_approved), 2),
            },
            "payroll_committed": round(payroll_committed, 2),
            "payroll_entries_count": len(payroll_entries),
        },
    }


# ─── Invoice Cost Update ─────────────────────────────────────────────────────

# VoR → tax jurisdiction mapping
_QUEBEC_VORS = {"digicom", "digicom inc", "digicom inc."}

def _calc_lender_tax(invoice):
    """Recalculate lender tax based on VoR province."""
    st = invoice.subtotal or invoice.total_due or 0
    margin = invoice.lender_margin_amt or 0
    base = st + margin
    vor = (invoice.vendor_on_record or "").strip().lower()
    if invoice.billing_type == "direct":
        return invoice.tax_total or 0  # pass through
    if vor in _QUEBEC_VORS:
        return round(base * 0.14975, 2)  # GST 5% + QST 9.975%
    return round(base * 0.13, 2)  # HST 13% (Ontario)


@router.put("/invoices/{invoice_id}/cost")
def update_invoice_cost(invoice_id: int, body: InvoiceCostUpdate, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    """Update billing/cost fields on an invoice."""
    inv = db.query(Invoice).filter(Invoice.id == invoice_id, Invoice.user_id == current_user.id).first()
    if not inv:
        raise HTTPException(status_code=404)
    allowed = {"lender_margin_pct", "govt_margin_pct", "lender_submitted_amt", "lender_approved_amt",
               "lender_status", "govt_submitted_amt", "govt_approved_amt", "govt_status"}
    data = body.model_dump(exclude_unset=True)
    # Validate margin percentages
    for pct_field in ("lender_margin_pct", "govt_margin_pct"):
        if pct_field in data and data[pct_field] is not None:
            if data[pct_field] < 0 or data[pct_field] > 200:
                raise HTTPException(status_code=400, detail=f"{pct_field} must be between 0 and 200")
    # Validate amounts are non-negative
    for amt_field in ("lender_submitted_amt", "lender_approved_amt", "govt_submitted_amt", "govt_approved_amt"):
        if amt_field in data and data[amt_field] is not None and data[amt_field] < 0:
            raise HTTPException(status_code=400, detail=f"{amt_field} cannot be negative")
    # Validate status values
    valid_statuses = {"pending", "approved", "partial", "rejected"}
    for status_field in ("lender_status", "govt_status"):
        if status_field in data and data[status_field] is not None and data[status_field] not in valid_statuses:
            raise HTTPException(status_code=400, detail=f"{status_field} must be one of: {', '.join(valid_statuses)}")
    for field, value in data.items():
        if field in allowed:
            setattr(inv, field, value)

    # Recalculate margins and tax
    st = inv.subtotal or inv.total_due or 0
    if inv.lender_margin_pct is not None:
        inv.lender_margin_amt = round(st * (inv.lender_margin_pct or 0) / 100, 2)
    if inv.govt_margin_pct is not None:
        inv.govt_margin_amt = round(st * (inv.govt_margin_pct or 0) / 100, 2)
    inv.received_total = inv.total_due
    inv.lender_tax_amt = _calc_lender_tax(inv)
    # Auto-calc submitted amounts if not manually set
    if inv.lender_submitted_amt is None:
        inv.lender_submitted_amt = round(st + (inv.lender_margin_amt or 0) + (inv.lender_tax_amt or 0), 2)
    if inv.govt_submitted_amt is None:
        inv.govt_submitted_amt = round(st + (inv.govt_margin_amt or 0), 2)
    db.commit()
    db.refresh(inv)
    return {"message": "Cost updated", "id": inv.id}


# ─── Payroll CRUD ────────────────────────────────────────────────────────────

def _calc_payroll(entry: PayrollEntry):
    """Calculate derived payroll fields."""
    entry.eligible_days = (entry.working_days or 0) - (entry.statutory_holidays or 0)
    if entry.eligible_days and entry.eligible_days > 0:
        entry.daily_rate = round((entry.gross_pay or 0) / entry.eligible_days, 2)
    else:
        entry.daily_rate = 0
    entry.lender_billable = entry.gross_pay  # lender approves full gross
    non_claimable = (entry.cpp or 0) + (entry.ei or 0) + (entry.insurance or 0) + (entry.holiday_pay or 0)
    entry.govt_billable = round((entry.gross_pay or 0) - non_claimable, 2)


@router.get("/payroll")
def list_payroll(proj: Optional[Project] = Depends(_get_proj), db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    q = db.query(PayrollEntry).filter(PayrollEntry.user_id == current_user.id)
    if proj:
        q = q.filter(PayrollEntry.project_id == proj.id)
    return [PayrollEntryOut.model_validate(e) for e in q.order_by(PayrollEntry.pay_period_start.desc()).all()]


def _validate_payroll(body):
    """Reject negative monetary values in payroll entries."""
    for field in ("gross_pay", "cpp", "ei", "income_tax", "insurance", "holiday_pay", "other_deductions"):
        val = getattr(body, field, None) if hasattr(body, field) else body.get(field) if isinstance(body, dict) else None
        if val is not None and val < 0:
            raise HTTPException(status_code=400, detail=f"{field} cannot be negative")


@router.post("/payroll")
def create_payroll(body: PayrollEntryCreate, proj: Project = Depends(_req_proj), db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    _validate_payroll(body)
    entry = PayrollEntry(user_id=current_user.id, project_id=proj.id, status="processed")
    for field, value in body.model_dump().items():
        if hasattr(entry, field):
            setattr(entry, field, value)
    _calc_payroll(entry)
    db.add(entry)
    db.commit()
    db.refresh(entry)
    return PayrollEntryOut.model_validate(entry)


@router.put("/payroll/{entry_id}")
def update_payroll(entry_id: int, body: PayrollEntryUpdate, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    _validate_payroll(body)
    entry = db.query(PayrollEntry).filter(PayrollEntry.id == entry_id, PayrollEntry.user_id == current_user.id).first()
    if not entry:
        raise HTTPException(status_code=404)
    allowed = {"employee_name", "company_name", "gross_pay", "cpp", "ei", "income_tax", "insurance",
               "holiday_pay", "working_days", "statutory_holidays", "province",
               "lender_submitted_amt", "lender_approved_amt", "lender_status",
               "govt_submitted_amt", "govt_approved_amt", "govt_status",
               "draw_id", "provincial_claim_id", "federal_claim_id"}
    for field, value in body.model_dump(exclude_unset=True).items():
        if field in allowed:
            setattr(entry, field, value)
    _calc_payroll(entry)
    db.commit()
    db.refresh(entry)
    return PayrollEntryOut.model_validate(entry)


@router.delete("/payroll/{entry_id}")
def delete_payroll(entry_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    entry = db.query(PayrollEntry).filter(PayrollEntry.id == entry_id, PayrollEntry.user_id == current_user.id).first()
    if not entry:
        raise HTTPException(status_code=404)
    db.delete(entry)
    db.commit()
    return {"message": "Payroll entry deleted"}


# ─── Bookkeeping Export ──────────────────────────────────────────────────────

def _auth_export(authorization: str, db: Session) -> User:
    from ..dependencies import SECRET_KEY, ALGORITHM
    from jose import jwt
    if not authorization or not authorization.startswith("Bearer "):
        raise ValueError("missing token")
    payload = jwt.decode(authorization[7:], SECRET_KEY, algorithms=[ALGORITHM])
    user = db.query(User).filter(User.id == int(payload.get("sub"))).first()
    if not user or not user.is_active:
        raise ValueError("invalid")
    return user


@router.get("/export/bookkeeping")
def export_bookkeeping(
    entity: Optional[str] = None,
    authorization: Optional[str] = Header(None),
    db: Session = Depends(get_db),
):
    """Export project finance data as Excel workbook for bookkeeping."""
    try:
        current_user = _auth_export(authorization, db)
    except Exception:
        raise HTTPException(status_code=401, detail="Invalid or missing token")

    proj = (db.query(Project).filter(Project.user_id == current_user.id)
            .order_by(Project.created_at).first())
    if not proj:
        raise HTTPException(status_code=404, detail="No project found")

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    alt_fill = PatternFill(start_color="F0F4F8", end_color="F0F4F8", fill_type="solid")

    wb = openpyxl.Workbook()

    # ── Sheet 1: Cost Summary ────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Cost Summary"
    headers = ["Category", "Budget", "Invoiced", "Paid", "Remaining", "% Used"]
    for i, h in enumerate(headers, 1):
        c = ws1.cell(row=1, column=i, value=h)
        c.font = header_font; c.fill = header_fill; c.alignment = Alignment(horizontal="center")

    cats = db.query(CostCategory).filter(CostCategory.project_id == proj.id).order_by(CostCategory.display_order).all()
    row = 2
    for cat in cats:
        invoiced = db.query(func.coalesce(func.sum(InvoiceAllocation.amount), 0.0)).filter(InvoiceAllocation.category_id == cat.id).scalar()
        paid = 0.0
        for a in db.query(InvoiceAllocation).filter(InvoiceAllocation.category_id == cat.id).all():
            inv = db.query(Invoice).filter(Invoice.id == a.invoice_id).first()
            if inv:
                paid += (inv.amount_paid or 0) * (a.percentage / 100)
        remaining = cat.budget - invoiced
        pct = (invoiced / cat.budget * 100) if cat.budget else 0
        ws1.cell(row=row, column=1, value=cat.name)
        ws1.cell(row=row, column=2, value=round(cat.budget, 2))
        ws1.cell(row=row, column=3, value=round(invoiced, 2))
        ws1.cell(row=row, column=4, value=round(paid, 2))
        ws1.cell(row=row, column=5, value=round(remaining, 2))
        ws1.cell(row=row, column=6, value=round(pct, 1))
        if row % 2 == 0:
            for col in range(1, 7):
                ws1.cell(row=row, column=col).fill = alt_fill
        row += 1
    for i in range(1, 7):
        ws1.column_dimensions[get_column_letter(i)].width = 20

    # ── Sheet 2: Invoice Register ────────────────────────────────────────────
    ws2 = wb.create_sheet("Invoice Register")
    inv_headers = ["Invoice #", "Date", "Due Date", "Vendor", "Billed To", "Billing Type",
                   "Vendor on Record", "Amount", "Paid", "Balance", "Payment Status",
                   "Cost Category", "Sub-Category", "Sub-Division", "Allocation %", "Aging Days"]
    for i, h in enumerate(inv_headers, 1):
        c = ws2.cell(row=1, column=i, value=h)
        c.font = header_font; c.fill = header_fill; c.alignment = Alignment(horizontal="center")

    invoices_q = db.query(Invoice).filter(
        Invoice.user_id == current_user.id, Invoice.status == "processed"
    )
    if entity:
        invoices_q = invoices_q.filter(Invoice.billed_to.ilike(f"%{entity}%"))
    invoices = invoices_q.order_by(Invoice.invoice_date.desc()).all()

    today_str = datetime.utcnow().strftime("%Y-%m-%d")
    row = 2
    for inv in invoices:
        allocs = db.query(InvoiceAllocation).filter(InvoiceAllocation.invoice_id == inv.id).all()
        if not allocs:
            allocs = [None]  # still show the invoice even if unallocated

        for alloc in allocs:
            due = inv.due_date or inv.invoice_date
            try:
                aging_days = (datetime.strptime(today_str, "%Y-%m-%d") - datetime.strptime(due, "%Y-%m-%d")).days if due else 0
            except ValueError:
                aging_days = 0

            ws2.cell(row=row, column=1, value=inv.invoice_number)
            ws2.cell(row=row, column=2, value=inv.invoice_date)
            ws2.cell(row=row, column=3, value=inv.due_date)
            ws2.cell(row=row, column=4, value=inv.vendor_name)
            ws2.cell(row=row, column=5, value=inv.billed_to)
            ws2.cell(row=row, column=6, value=inv.billing_type)
            ws2.cell(row=row, column=7, value=inv.vendor_on_record)
            ws2.cell(row=row, column=8, value=round(inv.total_due or 0, 2))
            ws2.cell(row=row, column=9, value=round(inv.amount_paid or 0, 2))
            ws2.cell(row=row, column=10, value=round((inv.total_due or 0) - (inv.amount_paid or 0), 2))
            ws2.cell(row=row, column=11, value=inv.payment_status or "unpaid")

            if alloc:
                ws2.cell(row=row, column=12, value=alloc.category.name if alloc.category else "")
                ws2.cell(row=row, column=13, value=alloc.sub_category.name if alloc.sub_category else "")
                ws2.cell(row=row, column=14, value=alloc.subdivision.name if alloc.subdivision else "")
                ws2.cell(row=row, column=15, value=alloc.percentage)
            else:
                ws2.cell(row=row, column=12, value="UNALLOCATED")

            ws2.cell(row=row, column=16, value=aging_days if inv.payment_status != "paid" else 0)
            row += 1

    for i in range(1, len(inv_headers) + 1):
        ws2.column_dimensions[get_column_letter(i)].width = 18

    # ── Sheet 3: Payments Ledger ─────────────────────────────────────────────
    ws3 = wb.create_sheet("Payments Ledger")
    pmt_headers = ["Payment Date", "Invoice #", "Vendor", "Billed To", "Amount", "Method", "Reference", "Notes"]
    for i, h in enumerate(pmt_headers, 1):
        c = ws3.cell(row=1, column=i, value=h)
        c.font = header_font; c.fill = header_fill; c.alignment = Alignment(horizontal="center")

    payments = (
        db.query(Payment)
        .join(Invoice)
        .filter(Invoice.user_id == current_user.id)
        .order_by(Payment.payment_date.desc())
        .all()
    )
    row = 2
    for pmt in payments:
        inv = db.query(Invoice).filter(Invoice.id == pmt.invoice_id).first()
        if entity and inv and inv.billed_to and entity.lower() not in inv.billed_to.lower():
            continue
        ws3.cell(row=row, column=1, value=pmt.payment_date)
        ws3.cell(row=row, column=2, value=inv.invoice_number if inv else "")
        ws3.cell(row=row, column=3, value=inv.vendor_name if inv else "")
        ws3.cell(row=row, column=4, value=inv.billed_to if inv else "")
        ws3.cell(row=row, column=5, value=round(pmt.amount, 2))
        ws3.cell(row=row, column=6, value=pmt.method)
        ws3.cell(row=row, column=7, value=pmt.reference)
        ws3.cell(row=row, column=8, value=pmt.notes)
        row += 1
    for i in range(1, len(pmt_headers) + 1):
        ws3.column_dimensions[get_column_letter(i)].width = 18

    # ── Sheet 4: Aging Report ────────────────────────────────────────────────
    ws4 = wb.create_sheet("Aging Report")
    aging_headers = ["Invoice #", "Vendor", "Billed To", "Invoice Date", "Due Date", "Total", "Paid", "Balance", "Days Overdue", "Bucket"]
    for i, h in enumerate(aging_headers, 1):
        c = ws4.cell(row=1, column=i, value=h)
        c.font = header_font; c.fill = header_fill; c.alignment = Alignment(horizontal="center")

    unpaid_invoices = db.query(Invoice).filter(
        Invoice.user_id == current_user.id,
        Invoice.status == "processed",
        Invoice.payment_status != "paid",
    ).order_by(Invoice.due_date).all()

    row = 2
    for inv in unpaid_invoices:
        if entity and inv.billed_to and entity.lower() not in inv.billed_to.lower():
            continue
        due = inv.due_date or inv.invoice_date
        try:
            days = (datetime.strptime(today_str, "%Y-%m-%d") - datetime.strptime(due, "%Y-%m-%d")).days if due else 0
        except ValueError:
            days = 0
        bucket = "Current" if days <= 30 else "31-60" if days <= 60 else "61-90" if days <= 90 else "90+"
        balance = (inv.total_due or 0) - (inv.amount_paid or 0)

        ws4.cell(row=row, column=1, value=inv.invoice_number)
        ws4.cell(row=row, column=2, value=inv.vendor_name)
        ws4.cell(row=row, column=3, value=inv.billed_to)
        ws4.cell(row=row, column=4, value=inv.invoice_date)
        ws4.cell(row=row, column=5, value=inv.due_date)
        ws4.cell(row=row, column=6, value=round(inv.total_due or 0, 2))
        ws4.cell(row=row, column=7, value=round(inv.amount_paid or 0, 2))
        ws4.cell(row=row, column=8, value=round(balance, 2))
        ws4.cell(row=row, column=9, value=days)
        ws4.cell(row=row, column=10, value=bucket)
        row += 1
    for i in range(1, len(aging_headers) + 1):
        ws4.column_dimensions[get_column_letter(i)].width = 18

    # ── Sheet 5: Entity Summary ──────────────────────────────────────────────
    ws5 = wb.create_sheet("Entity Summary")
    ent_headers = ["Entity (Billed To)", "# Invoices", "Total Invoiced", "Total Paid", "Outstanding", "Billing Type"]
    for i, h in enumerate(ent_headers, 1):
        c = ws5.cell(row=1, column=i, value=h)
        c.font = header_font; c.fill = header_fill; c.alignment = Alignment(horizontal="center")

    # Group by billed_to
    all_inv = db.query(Invoice).filter(
        Invoice.user_id == current_user.id, Invoice.status == "processed"
    ).all()
    entity_map = {}
    for inv in all_inv:
        ent = inv.billed_to or "Unknown"
        if ent not in entity_map:
            entity_map[ent] = {"count": 0, "invoiced": 0.0, "paid": 0.0, "billing_type": set()}
        entity_map[ent]["count"] += 1
        entity_map[ent]["invoiced"] += (inv.total_due or 0)
        entity_map[ent]["paid"] += (inv.amount_paid or 0)
        if inv.billing_type:
            entity_map[ent]["billing_type"].add(inv.billing_type)

    row = 2
    for ent_name, data in sorted(entity_map.items()):
        ws5.cell(row=row, column=1, value=ent_name)
        ws5.cell(row=row, column=2, value=data["count"])
        ws5.cell(row=row, column=3, value=round(data["invoiced"], 2))
        ws5.cell(row=row, column=4, value=round(data["paid"], 2))
        ws5.cell(row=row, column=5, value=round(data["invoiced"] - data["paid"], 2))
        ws5.cell(row=row, column=6, value=", ".join(data["billing_type"]) if data["billing_type"] else "")
        row += 1
    for i in range(1, len(ent_headers) + 1):
        ws5.column_dimensions[get_column_letter(i)].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"bookkeeping_{proj.name.replace(' ', '_')}_{today_str}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


# ═══════════════════════════════════════════════════════════════════════════════
# AI INTELLIGENCE ENDPOINTS
# ═══════════════════════════════════════════════════════════════════════════════

# ─── AI Feature 1: Invoice → Cost Code Mapper ────────────────────────────────

@router.post("/ai/suggest-allocation")
def ai_suggest_allocation(
    body: dict,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Use Gemini to suggest the best cost category for an invoice based on vendor/description."""
    invoice_id = body.get("invoice_id")
    project_id = body.get("project_id")
    if not invoice_id:
        raise HTTPException(status_code=400, detail="invoice_id required")

    inv = db.query(Invoice).filter(Invoice.id == invoice_id, Invoice.user_id == current_user.id).first()
    if not inv:
        raise HTTPException(status_code=404, detail="Invoice not found")

    proj = None
    if project_id:
        proj = db.query(Project).filter(Project.id == project_id, Project.user_id == current_user.id).first()
    if not proj:
        proj = db.query(Project).filter(Project.user_id == current_user.id).order_by(Project.created_at).first()
    if not proj:
        raise HTTPException(status_code=404, detail="No project found")

    categories = db.query(CostCategory).filter(CostCategory.project_id == proj.id).order_by(CostCategory.display_order).all()
    if not categories:
        raise HTTPException(status_code=400, detail="No cost categories defined for this project")

    from ..services.ai_project import suggest_allocation
    result = suggest_allocation(inv, categories, db)
    return result


# ─── AI Feature 1b: Bulk auto-suggest for unallocated invoices ────────────────

@router.post("/ai/bulk-suggest-allocations")
def ai_bulk_suggest(
    body: dict,
    proj: Project = Depends(_req_proj),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Suggest allocations for all unallocated project invoices (up to 20 at a time)."""
    categories = db.query(CostCategory).filter(CostCategory.project_id == proj.id).order_by(CostCategory.display_order).all()
    if not categories:
        raise HTTPException(status_code=400, detail="No cost categories defined for this project")

    # Find unallocated processed invoices for this project
    unallocated = db.query(Invoice).filter(
        Invoice.user_id == current_user.id,
        Invoice.project_id == proj.id,
        Invoice.status == "processed",
    ).all()
    unallocated = [i for i in unallocated
                   if db.query(InvoiceAllocation).filter(InvoiceAllocation.invoice_id == i.id).count() == 0][:20]

    from ..services.ai_project import suggest_allocation
    results = []
    for inv in unallocated:
        suggestion = suggest_allocation(inv, categories, db)
        suggestion["invoice_id"] = inv.id
        suggestion["vendor"] = inv.vendor_name
        suggestion["amount"] = inv.total_due
        results.append(suggestion)

    return {"suggestions": results, "count": len(results)}


# ─── AI Feature 2: Lien & Holdback Compliance Brain ──────────────────────────

@router.get("/ai/compliance")
def ai_compliance(
    proj: Optional[Project] = Depends(_get_proj),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Canadian construction compliance alerts: holdback timelines, lien windows, missing waivers."""
    if not proj:
        return {"alerts": [], "alert_count": 0}

    invoices = db.query(Invoice).filter(
        Invoice.user_id == current_user.id,
        Invoice.project_id == proj.id,
        Invoice.status == "processed",
    ).all()
    draws = db.query(Draw).filter(Draw.project_id == proj.id).all()
    lien_waivers = db.query(LienWaiver).filter(LienWaiver.project_id == proj.id).all()

    from ..services.ai_project import compliance_alerts
    return compliance_alerts(proj, invoices, draws, lien_waivers)


# ─── AI Feature 3: Cost Overrun Early Warning ─────────────────────────────────

@router.get("/ai/overrun-alerts")
def ai_overrun_alerts(
    proj: Optional[Project] = Depends(_get_proj),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Detect categories trending toward budget overrun based on spend velocity."""
    if not proj:
        return {"alerts": [], "alert_count": 0}

    categories = db.query(CostCategory).filter(CostCategory.project_id == proj.id).order_by(CostCategory.display_order).all()

    # Build allocations by category
    allocations_by_cat: dict = {}
    for cat in categories:
        total = db.query(func.coalesce(func.sum(InvoiceAllocation.amount), 0.0)).filter(
            InvoiceAllocation.category_id == cat.id
        ).scalar() or 0.0
        allocations_by_cat[cat.id] = float(total)

    # Approved change orders by category
    approved_cos = db.query(ChangeOrder).filter(
        ChangeOrder.project_id == proj.id, ChangeOrder.status == "approved"
    ).all()
    co_by_cat: dict = {}
    for co in approved_cos:
        if co.category_id:
            co_by_cat[co.category_id] = co_by_cat.get(co.category_id, 0.0) + co.amount

    from ..services.ai_project import overrun_alerts
    return overrun_alerts(proj, categories, allocations_by_cat, co_by_cat)


# ─── AI Feature 4: Draw Intelligence Engine ───────────────────────────────────

@router.get("/ai/draw-readiness/{draw_id}")
def ai_draw_readiness(
    draw_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Generate a draw submission readiness checklist."""
    draw = (db.query(Draw).join(Project, Project.id == Draw.project_id)
            .filter(Draw.id == draw_id, Project.user_id == current_user.id).first())
    if not draw:
        raise HTTPException(status_code=404, detail="Draw not found")

    invoices = db.query(Invoice).filter(Invoice.draw_id == draw_id, Invoice.user_id == current_user.id).all()
    proj = db.query(Project).filter(Project.id == draw.project_id).first()
    lien_waivers = db.query(LienWaiver).filter(LienWaiver.project_id == proj.id).all()
    subcontractors = db.query(Subcontractor).filter(Subcontractor.project_id == proj.id).all()
    documents = db.query(ProjectDocument).filter(ProjectDocument.project_id == proj.id).all()

    from ..services.ai_project import draw_readiness
    return draw_readiness(draw, invoices, lien_waivers, subcontractors, documents)


# ─── AI Feature 5: Cash Flow Reality Simulator ────────────────────────────────

@router.get("/ai/cashflow-scenarios")
def ai_cashflow_scenarios(
    delay_months: int = 0,
    cost_inflation_pct: float = 0.0,
    draw_delay_days: int = 0,
    proj: Optional[Project] = Depends(_get_proj),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Simulate cash flow under stress scenarios (delays, cost inflation, draw delays)."""
    if not proj:
        return {"base": [], "stressed": [], "summary": {}}

    # Get base cash flow months (reuse existing cash_flow logic inline)
    from collections import defaultdict
    from datetime import datetime as _dt

    spend: dict = defaultdict(float)
    paid_map: dict = defaultdict(float)
    receipts: dict = defaultdict(float)
    projected: dict = defaultdict(float)

    proj_draws_cf = db.query(Draw).filter(Draw.project_id == proj.id).all()
    proj_draw_ids_cf = [d.id for d in proj_draws_cf]
    from sqlalchemy import or_ as _or_cf
    _cf_conds = [Invoice.project_id == proj.id]
    if proj_draw_ids_cf:
        _cf_conds.append(Invoice.draw_id.in_(proj_draw_ids_cf))
    _cf_inv = db.query(Invoice).filter(
        Invoice.user_id == current_user.id,
        Invoice.status == "processed",
        _or_cf(*_cf_conds),
    ).all()
    for inv in _cf_inv:
        date_str = inv.invoice_date or (str(inv.processed_at)[:10] if inv.processed_at else None)
        if date_str and len(date_str) >= 7:
            spend[date_str[:7]] += inv.total_due or 0

    _cf_inv_ids = {i.id for i in _cf_inv}
    for pmt in db.query(Payment).join(Invoice, Invoice.id == Payment.invoice_id).filter(
        Invoice.user_id == current_user.id,
        Invoice.id.in_(_cf_inv_ids) if _cf_inv_ids else Invoice.id == -1,
    ).all():
        if pmt.payment_date and len(pmt.payment_date) >= 7:
            paid_map[pmt.payment_date[:7]] += pmt.amount or 0

    for draw in proj_draws_cf:
        if draw.submission_date and len(draw.submission_date) >= 7:
            m = draw.submission_date[:7]
            draw_total = db.query(func.coalesce(func.sum(Invoice.lender_approved_amt), 0.0)).filter(
                Invoice.draw_id == draw.id, Invoice.user_id == current_user.id
            ).scalar() or 0
            receipts[m] += draw_total

    today_m = _dt.utcnow().strftime("%Y-%m")
    for cc in db.query(CommittedCost).filter(CommittedCost.project_id == proj.id, CommittedCost.status == "active").all():
        if cc.expected_completion and len(cc.expected_completion) >= 7:
            m = cc.expected_completion[:7]
            if m >= today_m:
                remaining = cc.contract_amount - (cc.invoiced_to_date or 0)
                if remaining > 0:
                    projected[m] += remaining

    all_months = sorted(set(list(spend.keys()) + list(paid_map.keys()) + list(receipts.keys()) + list(projected.keys())))
    cumulative = 0.0
    base_months = []
    for m in all_months:
        s = round(spend.get(m, 0), 2)
        p = round(paid_map.get(m, 0), 2)
        r = round(receipts.get(m, 0), 2)
        pr = round(projected.get(m, 0), 2)
        net = round(r - s, 2)
        cumulative = round(cumulative + net, 2)
        base_months.append({"month": m, "invoiced": s, "paid": p, "draw_receipts": r,
                             "projected_spend": pr, "net": net, "cumulative": cumulative})

    # Clamp scenario params
    delay_months = max(0, min(12, delay_months))
    cost_inflation_pct = max(0.0, min(50.0, cost_inflation_pct))
    draw_delay_days = max(0, min(180, draw_delay_days))

    from ..services.ai_project import cashflow_scenarios
    return cashflow_scenarios(base_months, proj, delay_months, cost_inflation_pct, draw_delay_days)


# ─── AI Feature 6: Subcontractor Risk Scores ─────────────────────────────────

@router.get("/ai/subcontractor-risks")
def ai_subcontractor_risks(
    proj: Optional[Project] = Depends(_get_proj),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Score each subcontractor 0–100 based on compliance and payment history (lower = more risky)."""
    if not proj:
        return []

    subcontractors = db.query(Subcontractor).filter(Subcontractor.project_id == proj.id).order_by(Subcontractor.name).all()
    if not subcontractors:
        return []

    invoices = db.query(Invoice).filter(
        Invoice.user_id == current_user.id,
        Invoice.project_id == proj.id,
        Invoice.status == "processed",
    ).all()
    change_orders = db.query(ChangeOrder).filter(ChangeOrder.project_id == proj.id).all()
    lien_waivers = db.query(LienWaiver).filter(LienWaiver.project_id == proj.id).all()

    from ..services.ai_project import subcontractor_risk_scores
    return subcontractor_risk_scores(subcontractors, invoices, change_orders, lien_waivers)


# ─── AI Feature 7: Lender Behavior Model ─────────────────────────────────────

@router.get("/ai/lender-insights")
def ai_lender_insights(
    proj: Optional[Project] = Depends(_get_proj),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Detect lender rejection patterns and provide submission optimization tips."""
    if not proj:
        return {"patterns": [], "tips": [], "pattern_count": 0, "tip_count": 0}

    draws = db.query(Draw).filter(Draw.project_id == proj.id).order_by(Draw.draw_number).all()
    invoices = db.query(Invoice).filter(
        Invoice.user_id == current_user.id,
        Invoice.project_id == proj.id,
        Invoice.status == "processed",
    ).all()
    lien_waivers = db.query(LienWaiver).filter(LienWaiver.project_id == proj.id).all()
    documents = db.query(ProjectDocument).filter(ProjectDocument.project_id == proj.id).all()

    from ..services.ai_project import lender_insights
    return lender_insights(draws, invoices, lien_waivers, documents)


# ─── AI Insights — Consolidated endpoint (all 7 in one call) ─────────────────

@router.get("/ai/insights")
def ai_insights_all(
    proj: Optional[Project] = Depends(_get_proj),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Run all AI intelligence features for the project in a single call."""
    if not proj:
        return {"project": None}

    from ..services.ai_project import (
        compliance_alerts, overrun_alerts,
        subcontractor_risk_scores, lender_insights,
    )

    categories = db.query(CostCategory).filter(CostCategory.project_id == proj.id).order_by(CostCategory.display_order).all()
    draws = db.query(Draw).filter(Draw.project_id == proj.id).order_by(Draw.draw_number).all()
    invoices = db.query(Invoice).filter(
        Invoice.user_id == current_user.id,
        Invoice.project_id == proj.id,
        Invoice.status == "processed",
    ).all()
    lien_waivers = db.query(LienWaiver).filter(LienWaiver.project_id == proj.id).all()
    subcontractors = db.query(Subcontractor).filter(Subcontractor.project_id == proj.id).all()
    change_orders = db.query(ChangeOrder).filter(ChangeOrder.project_id == proj.id).all()
    documents = db.query(ProjectDocument).filter(ProjectDocument.project_id == proj.id).all()

    # Allocations by category
    allocs_by_cat: dict = {}
    for cat in categories:
        total = db.query(func.coalesce(func.sum(InvoiceAllocation.amount), 0.0)).filter(
            InvoiceAllocation.category_id == cat.id
        ).scalar() or 0.0
        allocs_by_cat[cat.id] = float(total)

    # Approved COs by category
    co_by_cat: dict = {}
    for co in change_orders:
        if co.status == "approved" and co.category_id:
            co_by_cat[co.category_id] = co_by_cat.get(co.category_id, 0.0) + co.amount

    # Unallocated invoice count
    unallocated_count = sum(
        1 for inv in invoices
        if db.query(InvoiceAllocation).filter(InvoiceAllocation.invoice_id == inv.id).count() == 0
    )

    compliance = compliance_alerts(proj, invoices, draws, lien_waivers)
    overruns = overrun_alerts(proj, categories, allocs_by_cat, co_by_cat)
    sub_risks = subcontractor_risk_scores(subcontractors, invoices, change_orders, lien_waivers)
    lender = lender_insights(draws, invoices, lien_waivers, documents)

    # High-level badge counts for sidebar
    total_alerts = (
        compliance["alert_count"]
        + overruns["alert_count"]
        + lender["pattern_count"]
        + sum(1 for s in sub_risks if s["risk_level"] in ("critical", "high"))
    )

    return {
        "project_id": proj.id,
        "project_name": proj.name,
        "total_alerts": total_alerts,
        "unallocated_invoices": unallocated_count,
        "compliance": compliance,
        "overruns": overruns,
        "subcontractor_risks": sub_risks,
        "lender_insights": lender,
    }
